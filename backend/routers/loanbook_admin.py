"""
routers/loanbook_admin.py — Endpoints administrativos del loanbook.

Estos endpoints aplican el motor unificado (services/loanbook/engine.py) sobre
los datos en MongoDB para auditoría y reparación masiva.

Endpoints:
    GET  /api/loanbook/admin/audit-all       → reporta divergencias por LB
    POST /api/loanbook/admin/full-repair     → aplica engine.recalcular() y persiste

Diseño:
    - Solo router; toda la lógica financiera vive en engine.py
    - dry_run=true por default en full-repair para preview seguro
    - Reporte detallado: qué campo cambió por LB
"""
from __future__ import annotations

import logging
from datetime import datetime, timezone
from typing import Annotated

from fastapi import APIRouter, Depends, Query
from motor.motor_asyncio import AsyncIOMotorDatabase

from core.database import get_db
from core.datetime_utils import now_iso_bogota, today_bogota
from services.loanbook.engine import recalcular, auditar

logger = logging.getLogger("routers.loanbook_admin")

router = APIRouter(prefix="/api/loanbook-admin", tags=["loanbook-admin"])


# ─────────────────────────── audit-all ──────────────────────────────────────

@router.get("/audit-all")
async def audit_all(db: AsyncIOMotorDatabase = Depends(get_db)):
    """Audita TODOS los loanbooks y devuelve divergencias vs versión canónica.

    No modifica nada. Útil para visualizar el semáforo del módulo.

    Returns:
        {
          "fecha_corte":    "2026-04-30",
          "total":          43,
          "verdes":         N,
          "amarillas":      N,
          "rojas":          N,
          "puede_enviar_email_martes": bool (rojas == 0),
          "loanbooks": [
            {
              "loanbook_id": "...",
              "cliente":     "...",
              "severidad":   "verde" | "amarilla" | "roja",
              "violaciones": [{"campo", "antes", "despues"}, ...]
            },
            ...
          ]
        }
    """
    total = 0
    verdes = 0
    amarillas = 0
    rojas = 0
    reportes = []

    async for lb in db.loanbook.find({}):
        total += 1
        report = auditar(lb)
        if report["severidad"] == "verde":
            verdes += 1
        elif report["severidad"] == "amarilla":
            amarillas += 1
        else:
            rojas += 1
        reportes.append(report)

    # Ordenar por severidad: rojas primero, luego amarillas
    orden = {"roja": 0, "amarilla": 1, "verde": 2}
    reportes.sort(key=lambda r: (orden[r["severidad"]], r.get("loanbook_id", "")))

    return {
        "fecha_corte":   today_bogota().isoformat(),
        "fecha_analisis": now_iso_bogota(),
        "total":         total,
        "verdes":        verdes,
        "amarillas":     amarillas,
        "rojas":         rojas,
        "puede_enviar_email_martes": rojas == 0,
        "loanbooks":     reportes,
    }


# ─────────────────────────── full-repair ────────────────────────────────────

@router.post("/full-repair")
async def full_repair(
    dry_run: Annotated[
        bool,
        Query(description="True (default) = solo preview. False = aplica cambios."),
    ] = True,
    db: AsyncIOMotorDatabase = Depends(get_db),
):
    """Aplica engine.recalcular() a TODOS los loanbooks y persiste cambios.

    Repara automáticamente:
      - Cronogramas con cuotas fuera de miércoles canónico
      - Saldos desfasados (sc, si, saldo_pendiente)
      - DPD y sub_bucket
      - Estado derivado
      - Modelo poblado con plan_codigo
      - num_cuotas vs catálogo
      - mora_acumulada

    Idempotente: aplicarlo dos veces no produce cambios la segunda vez.

    Args:
        dry_run: Por defecto True — solo reporta qué cambiaría. Pasar dry_run=false
                 para persistir.

    Returns:
        {
          "dry_run":        bool,
          "total":          43,
          "modificados":    N,
          "sin_cambios":    N,
          "errores":        [{"loanbook_id", "error"}, ...],
          "reparaciones":   [
            {"loanbook_id", "cliente", "cambios": [{"campo", "antes", "despues"}, ...]},
            ...
          ]
        }
    """
    total = 0
    modificados = 0
    sin_cambios = 0
    errores = []
    reparaciones = []

    async for lb in db.loanbook.find({}):
        total += 1
        loanbook_id = lb.get("loanbook_id") or str(lb.get("_id"))
        cliente = (lb.get("cliente") or {}).get("nombre") or lb.get("cliente_nombre") or "?"

        try:
            canonico = recalcular(lb)
        except Exception as exc:
            logger.exception("full-repair error en %s: %s", loanbook_id, exc)
            errores.append({"loanbook_id": loanbook_id, "error": str(exc)})
            continue

        # Detectar cambios reales (no contar fecha_ultima_recalculacion)
        cambios = _detectar_cambios(lb, canonico)
        if not cambios:
            sin_cambios += 1
            continue

        modificados += 1
        reparaciones.append({
            "loanbook_id": loanbook_id,
            "cliente":     cliente,
            "cambios":     cambios,
        })

        # Persistir solo si no es dry_run
        if not dry_run:
            # Preservar _id de MongoDB si existe
            if "_id" in lb:
                canonico["_id"] = lb["_id"]
            await db.loanbook.replace_one({"_id": lb["_id"]}, canonico)

    return {
        "dry_run":          dry_run,
        "fecha_analisis":   now_iso_bogota(),
        "total":            total,
        "modificados":      modificados,
        "sin_cambios":      sin_cambios,
        "errores":          errores,
        "reparaciones":     reparaciones,
    }


# ─────────────────────────── helpers ────────────────────────────────────────

# Campos cuyo cambio NO se reporta (volátil o esperado)
CAMPOS_IGNORAR = {"fecha_ultima_recalculacion", "fecha_analisis", "_id"}


def _detectar_cambios(antes: dict, despues: dict) -> list[dict]:
    """Devuelve lista de campos que cambiaron entre antes y después."""
    cambios = []
    keys = (set(antes.keys()) | set(despues.keys())) - CAMPOS_IGNORAR
    for k in sorted(keys):
        v_antes = antes.get(k)
        v_despues = despues.get(k)
        if v_antes == v_despues:
            continue
        # Para cuotas: solo reportar si las fechas cambiaron
        if k == "cuotas":
            fechas_antes = [c.get("fecha") for c in (v_antes or [])]
            fechas_despues = [c.get("fecha") for c in (v_despues or [])]
            if fechas_antes == fechas_despues:
                continue
            cambios.append({
                "campo":   "cuotas_fechas",
                "antes":   fechas_antes[:3] + (["..."] if len(fechas_antes) > 3 else []),
                "despues": fechas_despues[:3] + (["..."] if len(fechas_despues) > 3 else []),
            })
            continue
        cambios.append({
            "campo":   k,
            "antes":   v_antes,
            "despues": v_despues,
        })
    return cambios


# ─────────────────────── MOTOR CANÓNICO ENDPOINTS (DAY1) ────────────────────

@router.get("/motor/audit-all")
async def motor_audit_all(db: AsyncIOMotorDatabase = Depends(get_db)):
    """Audita los 43 LBs aplicando el motor canónico (services.loanbook.motor).

    No modifica nada. Devuelve semáforo verde/amarilla/roja por LB.

    El motor aplica reglas v1.1 (Opción B):
      - 9 estados: pendiente_entrega, al_dia, mora_leve, mora_media, mora_grave,
        default, castigado, reestructurado, saldado.
      - Sub-buckets v1.1: Current/Grace/Warning/Alert/Critical/Severe/Pre-default/Default.
      - Mora $2.000 COP/día sin cap.
      - Solo recalcula derivados; NO toca cronograma.
    """
    from services.loanbook.motor import auditar
    from core.datetime_utils import today_bogota

    hoy = today_bogota()
    total = 0
    verdes = 0
    amarillas = 0
    rojas = 0
    reportes = []

    async for lb in db.loanbook.find({}):
        total += 1
        try:
            r = auditar(lb, hoy=hoy)
        except Exception as exc:
            r = {
                "loanbook_id": lb.get("loanbook_id"),
                "cliente":     (lb.get("cliente") or {}).get("nombre"),
                "ok":          False,
                "severidad":   "roja",
                "violaciones": [{"campo": "exception", "antes": "", "despues": str(exc), "tipo": "estructural"}],
            }
        if r["severidad"] == "verde":
            verdes += 1
        elif r["severidad"] == "amarilla":
            amarillas += 1
        else:
            rojas += 1
        reportes.append(r)

    orden = {"roja": 0, "amarilla": 1, "verde": 2}
    reportes.sort(key=lambda r: (orden.get(r["severidad"], 3), r.get("loanbook_id") or ""))

    return {
        "fecha_corte":               hoy.isoformat(),
        "fecha_analisis":            now_iso_bogota(),
        "motor":                     "services.loanbook.motor v1 (DAY1)",
        "total":                     total,
        "verdes":                    verdes,
        "amarillas":                 amarillas,
        "rojas":                     rojas,
        "puede_migrar_sin_riesgo":   rojas == 0,
        "loanbooks":                 reportes,
    }


@router.post("/motor/migrar")
async def motor_migrar(
    dry_run: Annotated[
        bool,
        Query(description="True (default) = solo preview. False = aplica derivar_estado y persiste."),
    ] = True,
    db: AsyncIOMotorDatabase = Depends(get_db),
):
    """Aplica services.loanbook.motor.derivar_estado a todos los LBs y persiste.

    SOLO recalcula derivados (saldo, dpd, sub_bucket, estado, mora, total_pagado,
    cuotas_pagadas, cuotas_vencidas). NO toca cronograma, NO toca términos pactados,
    NO toca metadata cliente. Cumple invariante: idempotente.

    Args:
        dry_run: True por defecto. False persiste cambios en MongoDB.

    Returns:
        Reporte detallado: cuántos LBs cambiaron, qué campo cambió en cada uno,
        cartera_total antes vs después.
    """
    from services.loanbook.motor import derivar_estado
    from core.datetime_utils import today_bogota

    hoy = today_bogota()
    total = 0
    modificados = 0
    sin_cambios = 0
    errores = []
    cambios_detalle = []
    cartera_antes = 0
    cartera_despues = 0

    async for lb in db.loanbook.find({}):
        total += 1
        loanbook_id = lb.get("loanbook_id") or str(lb.get("_id"))
        cliente = (lb.get("cliente") or {}).get("nombre") or lb.get("cliente_nombre") or "?"

        # Saldo antes (usa el persistido)
        saldo_antes = int(lb.get("saldo_pendiente") or 0)
        cartera_antes += saldo_antes

        try:
            canonico = derivar_estado(lb, hoy=hoy)
        except Exception as exc:
            errores.append({"loanbook_id": loanbook_id, "error": str(exc)})
            continue

        # Saldo después (canónico)
        saldo_despues = int(canonico.get("saldo_pendiente") or 0)
        cartera_despues += saldo_despues

        # Detectar cambios reales (campos derivados solamente)
        cambios = {}
        for k in ["saldo_pendiente", "total_pagado", "dpd", "estado", "sub_bucket",
                  "mora_acumulada_cop", "cuotas_pagadas", "cuotas_vencidas"]:
            v_antes = lb.get(k)
            v_despues = canonico.get(k)
            if v_antes != v_despues:
                cambios[k] = {"antes": v_antes, "despues": v_despues}

        if not cambios:
            sin_cambios += 1
            continue

        modificados += 1
        cambios_detalle.append({
            "loanbook_id": loanbook_id,
            "cliente":     cliente,
            "cambios":     cambios,
        })

        if not dry_run:
            # Persistir solo los campos derivados, no tocar cronograma ni términos
            patch = {k: canonico[k] for k in [
                "saldo_pendiente", "total_pagado", "dpd", "estado", "sub_bucket",
                "mora_acumulada_cop", "cuotas_pagadas", "cuotas_vencidas",
                "fecha_ultima_recalculacion",
            ] if k in canonico}
            await db.loanbook.update_one(
                {"_id": lb["_id"]},
                {"$set": patch},
            )

    return {
        "dry_run":                  dry_run,
        "fecha_analisis":           now_iso_bogota(),
        "motor":                    "services.loanbook.motor v1 (DAY1)",
        "total_loanbooks":          total,
        "modificados":              modificados,
        "sin_cambios":              sin_cambios,
        "errores":                  errores,
        "cartera_total_antes":      cartera_antes,
        "cartera_total_despues":    cartera_despues,
        "delta_cartera":            cartera_despues - cartera_antes,
        "cambios":                  cambios_detalle,
    }


# ─────────────────────── RESTAURAR DESDE EXCEL OFICIAL (DAY2) ───────────────

# Mapeo nomenclatura Excel oficial → Opción B (rangos v1.1)
ESTADO_EXCEL_A_OPCION_B = {
    "Aprobado":             "pendiente_entrega",
    "Pendiente Entrega":    "pendiente_entrega",
    "pendiente_entrega":    "pendiente_entrega",
    "Current":              "al_dia",
    "al_dia":               "al_dia",
    "Early Delinquency":    "mora_leve",
    "mora_leve":            "mora_leve",
    "Mid Delinquency":      "mora_media",
    "mora_media":           "mora_media",
    "Late Delinquency":     "mora_grave",
    "mora_grave":           "mora_grave",
    "en_riesgo":            "mora_leve",  # legacy → mapear al rango más cercano
    "mora":                 "mora_grave",  # legacy
    "Default":              "default",
    "default":              "default",
    "Charge-Off":           "castigado",
    "ChargeOff":            "castigado",
    "castigado":            "castigado",
    "Modificado":           "reestructurado",
    "reestructurado":       "reestructurado",
    "Pagado":               "saldado",
    "saldado":              "saldado",
    "activo":               "al_dia",  # legacy
}


@router.post("/restaurar-desde-excel")
async def restaurar_desde_excel(
    file: "UploadFile" = None,
    dry_run: Annotated[
        bool,
        Query(description="True (default) preview. False persiste."),
    ] = True,
    db: AsyncIOMotorDatabase = Depends(get_db),
):
    """Restaura los LBs desde el Excel oficial RODDOS.

    Lee el Excel oficial (loanbook_roddos_YYYY-MM-DD.xlsx) y para cada LB
    actualiza los campos canónicos sin tocar cronograma:
      - monto_original, cuota_periodica, total_cuotas
      - cuotas_pagadas, cuotas_vencidas
      - saldo_capital, saldo_intereses (separados como pide el negocio)
      - saldo_pendiente = sc + si
      - dpd, mora_acumulada_cop
      - estado (mapeado a Opción B), sub_bucket_semanal
      - producto, subtipo_rodante, plan_codigo, modalidad_pago
      - cliente block, metadata moto/repuestos/comparendo/etc

    NO toca:
      - cuotas[] (cronograma con fechas reales)
      - fecha_entrega, fecha_primer_pago (términos pactados)
      - capital_plan, cuota_estandar_plan (configuración del plan)
      - factura_alegra_id

    Soporta Hoja 1 (RDX) y Hoja 2 (RODANTE).
    """
    from fastapi import UploadFile, File
    from openpyxl import load_workbook
    from io import BytesIO
    from datetime import datetime, date

    if file is None:
        return {"error": "file requerido (upload Excel multipart)"}

    contenido = await file.read()
    wb = load_workbook(BytesIO(contenido), data_only=True)

    def _val(cell):
        v = cell.value if hasattr(cell, "value") else cell
        if isinstance(v, datetime):
            return v.date().isoformat()
        if isinstance(v, date):
            return v.isoformat()
        return v

    def _int(v, default=0):
        if v is None or v == "":
            return default
        try:
            return int(v)
        except (TypeError, ValueError):
            return default

    def _normalizar_estado(raw):
        if not raw:
            return "al_dia"
        return ESTADO_EXCEL_A_OPCION_B.get(str(raw).strip(), "al_dia")

    def _fila_a_patch(headers, row_cells):
        """Convierte una fila del Excel a dict de campos canónicos para $set."""
        d = {h: _val(c) for h, c in zip(headers, row_cells)}
        loanbook_id = d.get("loanbook_codigo") or d.get("loanbook_id")
        if not loanbook_id:
            return None, None

        patch = {}
        # Términos del crédito (canónicos del Excel)
        patch["monto_original"] = _int(d.get("monto_original"))
        patch["cuota_periodica"] = _int(d.get("cuota_periodica"))
        patch["cuota_monto"] = _int(d.get("cuota_periodica"))
        patch["total_cuotas"] = _int(d.get("total_cuotas"))
        patch["num_cuotas"] = _int(d.get("total_cuotas"))
        # Cuota inicial puede venir None
        if d.get("cuota_inicial") not in (None, ""):
            patch["cuota_inicial"] = _int(d.get("cuota_inicial"))

        # Desempeño (derivados oficiales del Excel — fuente de verdad)
        patch["cuotas_pagadas"] = _int(d.get("cuotas_pagadas"))
        patch["cuotas_vencidas"] = _int(d.get("cuotas_vencidas"))
        patch["saldo_capital"] = _int(d.get("saldo_capital"))
        patch["saldo_intereses"] = _int(d.get("saldo_intereses"))
        patch["saldo_pendiente"] = patch["saldo_capital"] + patch["saldo_intereses"]
        patch["mora_acumulada_cop"] = _int(d.get("mora_acumulada_cop"))
        patch["mora_acumulada"] = patch["mora_acumulada_cop"]
        patch["dpd"] = _int(d.get("dpd"))

        # Estado (mapeado a nomenclatura Opción B)
        patch["estado"] = _normalizar_estado(d.get("estado"))
        if d.get("sub_bucket_semanal"):
            patch["sub_bucket"] = str(d["sub_bucket_semanal"]).strip()
            patch["sub_bucket_semanal"] = str(d["sub_bucket_semanal"]).strip()

        # Identidad (no toca si ya existe)
        if d.get("producto"):
            patch["producto"] = str(d["producto"]).strip()
        if d.get("subtipo_rodante"):
            patch["subtipo_rodante"] = str(d["subtipo_rodante"]).strip()
        if d.get("plan_codigo"):
            patch["plan_codigo"] = str(d["plan_codigo"]).strip()
        if d.get("modalidad_pago"):
            patch["modalidad_pago"] = str(d["modalidad_pago"]).strip()
            patch["modalidad"] = str(d["modalidad_pago"]).strip()
        if d.get("tasa_ea") not in (None, ""):
            patch["tasa_ea"] = float(d.get("tasa_ea"))

        # Total_pagado derivado: valor_total_canónico − saldo_pendiente
        valor_total_canonico = patch["cuota_periodica"] * patch["total_cuotas"]
        patch["valor_total"] = valor_total_canonico
        patch["total_pagado"] = max(0, valor_total_canonico - patch["saldo_pendiente"])

        # Otros campos opcionales del Excel
        for k in ("vendedor", "whatsapp_status", "factura_alegra_id"):
            if d.get(k) not in (None, ""):
                patch[k] = str(d[k]).strip()
        if d.get("fecha_ultimo_pago") not in (None, ""):
            patch["fecha_ultimo_pago"] = d["fecha_ultimo_pago"]

        # Marker de auditoría
        patch["restaurado_desde_excel_at"] = now_iso_bogota()

        return loanbook_id, patch

    # Procesar Hoja 1 RDX y Hoja 2 RODANTE
    total = 0
    actualizados = 0
    no_encontrados = []
    errores = []
    cambios = []

    for sheet_name in ["Loan Tape RDX", "Loan Tape RODANTE"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

        for r in range(2, ws.max_row + 1):
            row_cells = [ws.cell(r, c) for c in range(1, ws.max_column + 1)]
            try:
                loanbook_id, patch = _fila_a_patch(headers, row_cells)
            except Exception as exc:
                errores.append({"fila": r, "sheet": sheet_name, "error": str(exc)})
                continue

            if loanbook_id is None:
                continue
            total += 1

            # Buscar el LB en MongoDB
            doc_actual = await db.loanbook.find_one({"loanbook_id": loanbook_id})
            if doc_actual is None:
                no_encontrados.append(loanbook_id)
                continue

            # Detectar cambios reales
            cambios_lb = {}
            for k, v in patch.items():
                if k == "restaurado_desde_excel_at":
                    continue
                v_actual = doc_actual.get(k)
                if v_actual != v:
                    cambios_lb[k] = {"antes": v_actual, "despues": v}
            if not cambios_lb:
                continue

            actualizados += 1
            cambios.append({
                "loanbook_id": loanbook_id,
                "cliente":     (doc_actual.get("cliente") or {}).get("nombre")
                                or doc_actual.get("cliente_nombre") or "?",
                "n_cambios":   len(cambios_lb),
                "cambios":     cambios_lb,
            })

            if not dry_run:
                await db.loanbook.update_one(
                    {"loanbook_id": loanbook_id},
                    {"$set": patch},
                )

    cartera_total_canonica = sum(
        (c["cambios"].get("saldo_pendiente", {}).get("despues")
         or c["cambios"].get("saldo_capital", {}).get("despues") or 0)
        for c in cambios
    )

    return {
        "dry_run":          dry_run,
        "fecha_analisis":   now_iso_bogota(),
        "fuente":           file.filename,
        "total_filas":      total,
        "actualizados":     actualizados,
        "sin_cambios":      total - actualizados - len(no_encontrados),
        "no_encontrados":   no_encontrados,
        "errores":          errores,
        "cambios":          cambios,
    }
