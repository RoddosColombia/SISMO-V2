"""
Bank statement parsers — 5 banks + auto-detect.

Supported formats:
- Bancolombia: .xlsx, sheet "Extracto", headers row 15
- BBVA: .xlsx, headers row 14
- Davivienda: .xlsx, skiprows=4
- Nequi: PDF via pdfplumber
- Global66: .xls/.xlsx, sheet "Movimientos de cuenta COP", headers row 4

Each parser returns: list[dict] with keys: fecha, descripcion, monto, tipo, banco
- fecha: yyyy-MM-dd format
- monto: absolute value in COP (always positive)
- tipo: "debito" or "credito"
"""
import hashlib
import os
import re
import shutil
import tempfile
from datetime import datetime

import openpyxl
import pandas as pd
import pdfplumber
from io import BytesIO


def detect_bank(file_path: str) -> str:
    """Auto-detect bank from file extension and header patterns."""
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".pdf":
        return "nequi"

    if ext not in (".xlsx", ".xls"):
        raise ValueError(f"Formato no soportado: {ext}. Solo .xlsx y .pdf.")

    wb = _open_workbook(file_path)

    # Global66: sheet "Movimientos de cuenta COP"
    if any("Movimientos de cuenta COP" in name for name in wb.sheetnames):
        wb.close()
        return "global66"

    # Bancolombia: sheet named "Extracto", headers at row 15
    if "Extracto" in wb.sheetnames:
        wb.close()
        return "bancolombia"

    ws = wb.active

    # Nequi xlsx: scan first 10 rows for Fecha+Descripcion+Valor headers
    for row_num in range(1, 11):
        try:
            row_cells = [str(c.value or "").strip().lower() for c in ws[row_num]]
            has_fecha = any("fecha" in c for c in row_cells)
            has_valor = any(c in ("valor", "monto", "transacción", "transaccion") for c in row_cells)
            if has_fecha and has_valor:
                wb.close()
                return "nequi"
        except Exception:
            continue

    # BBVA: headers at row 14 with "FECHA DE OPERACIÓN"
    row14 = [str(c.value or "") for c in ws[14]]
    if any("FECHA DE OPERACI" in cell.upper() for cell in row14):
        wb.close()
        return "bbva"

    # Davivienda: headers around row 5 with "Naturaleza" column
    for row_num in range(4, 7):
        row_cells = [str(c.value or "") for c in ws[row_num]]
        if any("naturaleza" in cell.lower() for cell in row_cells):
            wb.close()
            return "davivienda"

    wb.close()
    raise ValueError("No se pudo identificar el banco del extracto. Formatos soportados: Bancolombia, BBVA, Davivienda, Global66 (.xlsx/.xls) y Nequi (.xlsx o .pdf).")


def _open_workbook(file_path: str):
    """Open workbook, handling .xls files that are internally xlsx."""
    try:
        return openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception:
        # .xls extension may confuse openpyxl — copy to temp .xlsx and retry
        tmp_dir = tempfile.mkdtemp()
        tmp_path = os.path.join(tmp_dir, "extract.xlsx")
        shutil.copy2(file_path, tmp_path)
        return openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)


def _parse_date(date_str: str, fmt: str) -> str:
    """Parse date string to yyyy-MM-dd format."""
    try:
        if isinstance(date_str, datetime):
            return date_str.strftime("%Y-%m-%d")
        dt = datetime.strptime(str(date_str).strip(), fmt)
        return dt.strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return str(date_str).strip()


def _parse_monto(valor) -> tuple[float, str]:
    """Parse monetary value. Returns (abs_monto, tipo)."""
    if isinstance(valor, (int, float)):
        v = float(valor)
        return abs(v), "credito" if v >= 0 else "debito"

    s = str(valor).strip().replace("$", "").replace(" ", "")
    # Handle comma as thousands, dot as decimal: $-2,919.54 or $250,000.00
    negative = s.startswith("-")
    s = s.lstrip("-")
    s = s.replace(",", "")  # remove thousands separator
    try:
        v = float(s)
    except ValueError:
        v = 0.0
    tipo = "debito" if negative else "credito"
    return abs(v), tipo


def parse_bancolombia(file_path: str) -> list[dict]:
    """Bancolombia: sheet 'Extracto', headers row 15, cols FECHA/DESCRIPCION/VALOR."""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb["Extracto"]

    # Headers at row 15
    headers = [str(c.value or "").strip().upper() for c in ws[15]]
    col_map = {}
    for i, h in enumerate(headers):
        if "FECHA" in h:
            col_map["fecha"] = i
        elif "DESCRIPCI" in h:
            col_map["descripcion"] = i
        elif "VALOR" in h:
            col_map["valor"] = i

    movements = []
    for row in ws.iter_rows(min_row=16, values_only=True):
        if not row or not row[col_map.get("fecha", 0)]:
            continue
        fecha_raw = row[col_map["fecha"]]
        desc = str(row[col_map.get("descripcion", 1)] or "")
        valor_raw = row[col_map.get("valor", 2)]

        fecha = _parse_date(str(fecha_raw), "%d/%m")
        # Bancolombia d/m format lacks year — assume current year
        if len(fecha) <= 5:
            fecha = f"{datetime.now().year}-{fecha}"

        monto, tipo = _parse_monto(valor_raw)
        if monto == 0:
            continue

        movements.append({
            "fecha": fecha,
            "descripcion": desc.strip(),
            "monto": monto,
            "tipo": tipo,
            "banco": "Bancolombia",
        })

    wb.close()
    return movements


def parse_bbva(file_path: str) -> list[dict]:
    """BBVA: headers row 14, cols FECHA DE OPERACIÓN/CONCEPTO/IMPORTE."""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    headers = [str(c.value or "").strip().upper() for c in ws[14]]
    col_map = {}
    for i, h in enumerate(headers):
        if "FECHA" in h and "OPERACI" in h:
            col_map["fecha"] = i
        elif "CONCEPTO" in h:
            col_map["descripcion"] = i
        elif "IMPORTE" in h:
            col_map["valor"] = i

    movements = []
    for row in ws.iter_rows(min_row=15, values_only=True):
        if not row or not row[col_map.get("fecha", 0)]:
            continue
        fecha_raw = row[col_map["fecha"]]
        desc = str(row[col_map.get("descripcion", 1)] or "")
        valor_raw = row[col_map.get("valor", 2)]

        fecha = _parse_date(str(fecha_raw), "%d-%m-%Y")
        monto, tipo = _parse_monto(valor_raw)
        if monto == 0:
            continue

        movements.append({
            "fecha": fecha,
            "descripcion": desc.strip(),
            "monto": monto,
            "tipo": tipo,
            "banco": "BBVA",
        })

    wb.close()
    return movements


def parse_davivienda(file_path: str) -> list[dict]:
    """Davivienda: skiprows=4, cols Fecha/Descripcion/Valor/Naturaleza (C/D)."""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    # Headers at row 5 (after skipping 4)
    headers = [str(c.value or "").strip().upper() for c in ws[5]]
    col_map = {}
    for i, h in enumerate(headers):
        if "FECHA" in h:
            col_map["fecha"] = i
        elif "DESCRIPCI" in h:
            col_map["descripcion"] = i
        elif "VALOR" in h:
            col_map["valor"] = i
        elif "NATURALEZA" in h:
            col_map["naturaleza"] = i

    movements = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        if not row or not row[col_map.get("fecha", 0)]:
            continue
        fecha_raw = row[col_map["fecha"]]
        desc = str(row[col_map.get("descripcion", 1)] or "")
        valor_raw = row[col_map.get("valor", 2)]
        naturaleza = str(row[col_map.get("naturaleza", 3)] or "").strip().upper()

        fecha = _parse_date(str(fecha_raw), "%Y-%m-%d")
        monto, _ = _parse_monto(valor_raw)
        tipo = "credito" if naturaleza == "C" else "debito"
        if monto == 0:
            continue

        movements.append({
            "fecha": fecha,
            "descripcion": desc.strip(),
            "monto": monto,
            "tipo": tipo,
            "banco": "Davivienda",
        })

    wb.close()
    return movements


def parse_nequi(file_path: str, password: str | None = None) -> list[dict]:
    """Nequi: xlsx (preferred) or PDF via pdfplumber.

    xlsx columns: Fecha | Descripcion (or similar) | Valor | Saldo
    Negative valor = egreso (debito), positive = ingreso (credito).

    Args:
        file_path: Path to .xlsx or .pdf extract
        password: PDF password only (Nequi uses cédula del titular)
    """
    ext = os.path.splitext(file_path)[1].lower()
    if ext in (".xlsx", ".xls"):
        return _parse_nequi_xlsx(file_path)
    return _parse_nequi_pdf(file_path, password)


def _parse_nequi_xlsx(file_path: str) -> list[dict]:
    """Parse Nequi xlsx — pandas with multi-sheet and multi-column fallback (ported from V1)."""
    # Sheet name candidates
    SHEETS = ["Extracto Nequi", "Extracto", "Movimientos", "Hoja1", "Sheet1", 0]
    COLS_FECHA = ["Fecha", "FECHA", "Date", "Fecha de operación", "Fecha Operación"]
    COLS_DESC  = ["Descripción", "DESCRIPCIÓN", "Descripcion", "DESCRIPCION",
                  "Concepto", "CONCEPTO", "Detalle", "Tipo de transacción"]
    COLS_VALOR = ["Monto", "MONTO", "Valor", "VALOR", "Importe", "Amount"]
    COLS_TIPO  = ["Tipo", "TIPO", "Naturaleza", "Tipo de transacción",
                  "Tipo Transacción", "Dirección"]

    def _find_col(df_cols, options):
        upper_map = {c.upper(): c for c in df_cols}
        for opt in options:
            if opt in df_cols:
                return opt
            if opt.upper() in upper_map:
                return upper_map[opt.upper()]
        return None

    with open(file_path, "rb") as f:
        file_bytes = f.read()

    df = None
    for sheet in SHEETS:
        try:
            candidate = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet)
            if not candidate.empty:
                df = candidate
                break
        except Exception:
            continue

    if df is None or df.empty:
        raise ValueError("Nequi xlsx: no se pudo leer el extracto. Verifica que el archivo tiene datos.")

    col_fecha = _find_col(df.columns, COLS_FECHA)
    col_desc  = _find_col(df.columns, COLS_DESC)
    col_valor = _find_col(df.columns, COLS_VALOR)
    col_tipo  = _find_col(df.columns, COLS_TIPO)

    if not col_fecha or not col_valor:
        raise ValueError(f"Nequi xlsx: columnas requeridas no encontradas. Disponibles: {list(df.columns)}")

    movements = []
    for _, row in df.iterrows():
        try:
            fecha_raw = row[col_fecha]
            if pd.isna(fecha_raw):
                continue
            fecha = pd.to_datetime(fecha_raw).strftime("%Y-%m-%d")

            desc = (
                str(row[col_desc]).strip()
                if col_desc and not pd.isna(row.get(col_desc, float("nan")))
                else "Movimiento Nequi"
            )

            monto_raw = float(row[col_valor])
            if pd.isna(monto_raw):
                continue

            # Use Tipo column if available, else infer from sign
            if col_tipo and not pd.isna(row.get(col_tipo, float("nan"))):
                tipo_str = str(row[col_tipo]).strip().lower()
                tipo = "credito" if any(kw in tipo_str for kw in ["ingreso", "entrada", "abono", "recibo"]) else "debito"
            else:
                tipo = "credito" if monto_raw > 0 else "debito"

            monto = abs(monto_raw)
            if monto == 0:
                continue

            movements.append({
                "fecha": fecha,
                "descripcion": desc,
                "monto": monto,
                "tipo": tipo,
                "banco": "Nequi",
            })
        except Exception:
            continue

    return movements


def _parse_nequi_pdf(file_path: str, password: str | None = None) -> list[dict]:
    """Parse Nequi PDF extract (legacy path)."""
    movements = []
    open_kwargs: dict = {}
    if password:
        open_kwargs["password"] = password

    with pdfplumber.open(file_path, **open_kwargs) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    if not row or len(row) < 3:
                        continue
                    if row[0] and "fecha" in str(row[0]).lower():
                        continue

                    fecha_raw = str(row[0] or "").strip()
                    desc = str(row[1] or "").strip()
                    valor_raw = str(row[2] or "").strip()

                    if not fecha_raw or not valor_raw:
                        continue

                    fecha = _parse_date(fecha_raw, "%d/%m/%Y")
                    monto, tipo = _parse_monto(valor_raw)
                    if monto == 0:
                        continue

                    movements.append({
                        "fecha": fecha,
                        "descripcion": desc,
                        "monto": monto,
                        "tipo": tipo,
                        "banco": "Nequi",
                    })

    return movements


def parse_global66(file_path: str) -> list[dict]:
    """Global66: sheet 'Movimientos de cuenta COP', headers row 4, data row 5+.

    Columns (0-indexed):
    A(0): Tipo transaccion  B(1): Fecha  C(2): Monto debitado  D(3): Monto acreditado
    H(7): Nombre tercero  I(8): DNI tercero  M(12): ID transaccion  N(13): Comentario
    """
    wb = _open_workbook(file_path)

    # Find the target sheet
    ws = None
    for name in wb.sheetnames:
        if "Movimientos de cuenta COP" in name:
            ws = wb[name]
            break
    if ws is None:
        wb.close()
        raise ValueError("No se encontró hoja 'Movimientos de cuenta COP' en el archivo Global66.")

    movements = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row or len(row) < 14:
            continue

        tipo_tx = str(row[0] or "").strip()
        fecha_raw = str(row[1] or "").strip()
        monto_debito = row[2]
        monto_credito = row[3]
        nombre_tercero = str(row[7] or "").strip()
        referencia_id = str(row[12] or "").strip()
        comentario = str(row[13] or "").strip()

        # Determine tipo and monto
        has_debito = monto_debito is not None and monto_debito != "" and monto_debito != 0
        has_credito = monto_credito is not None and monto_credito != "" and monto_credito != 0

        if has_debito:
            monto_val, _ = _parse_monto(monto_debito)
            tipo = "debito"
        elif has_credito:
            monto_val, _ = _parse_monto(monto_credito)
            tipo = "credito"
        else:
            continue  # Skip rows with no value

        if monto_val == 0:
            continue

        # Parse fecha: YYYY-MM-DD HH:MM:SS → yyyy-MM-dd
        fecha = _parse_date(fecha_raw, "%Y-%m-%d %H:%M:%S")

        # Build enriched description
        parts = [p for p in [tipo_tx, comentario, nombre_tercero] if p]
        descripcion = " — ".join(parts) if parts else tipo_tx

        movements.append({
            "fecha": fecha,
            "descripcion": descripcion,
            "monto": monto_val,
            "tipo": tipo,
            "banco": "Global66",
            "referencia": referencia_id,
        })

    wb.close()
    return movements
