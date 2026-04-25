"""
services/loanbook/excel_export.py — Export de portafolio a Excel (.xlsx).

Produce un workbook con 2 hojas:
  - "Creditos": una fila por loanbook con comparación DB vs tabla PLAN_CUOTAS,
    campos financieros completos (saldo_intereses, cartera_total, DPD, mora)
    y datos de contacto del cliente.
  - "Cuotas": una fila por cuota con flags de corrupción (es_cuota_corrupta,
    motivo_corrupcion).

Función pura: recibe los loanbooks ya fetcheados, devuelve bytes del .xlsx.
Sin I/O de DB — el caller (endpoint HTTP) es responsable de traer los docs.
"""

from __future__ import annotations

import io
from datetime import date
from core.datetime_utils import now_bogota, today_bogota, now_iso_bogota

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from services.loanbook.reglas_negocio import get_num_cuotas, get_valor_total

# ─────────────────────── Paleta de estilos ────────────────────────────────────

_ROJO_FILL   = PatternFill("solid", fgColor="FFCCCC")   # diferencia detectada
_VERDE_FILL  = PatternFill("solid", fgColor="CCFFCC")   # valor correcto
_HEADER_FILL = PatternFill("solid", fgColor="1F497D")   # azul Roddos oscuro
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT        = Alignment(horizontal="left",   vertical="center")


def _header_row(ws, cols: list[str]) -> None:
    """Escribe la fila de cabecera con estilo."""
    ws.append(cols)
    for cell in ws[1]:
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _CENTER
    ws.row_dimensions[1].height = 28


def _autofit(ws, min_width: int = 10, max_width: int = 40) -> None:
    """Ajuste aproximado de ancho de columnas."""
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max_width, max(min_width, max_len + 2)
        )


def _cuota_corrupta(c: dict, hoy_str: str) -> tuple[bool, str]:
    """Determina si una cuota es corrupta y el motivo.

    Returns:
        (es_corrupta: bool, motivo: str)
    """
    estado   = c.get("estado", "")
    fecha_c  = c.get("fecha", "") or ""
    fecha_p  = c.get("fecha_pago", "") or ""
    ref      = c.get("referencia")
    metodo   = c.get("metodo_pago")
    tiene_ev = bool(ref) or bool(metodo and metodo not in ("", "seed", None))

    # Caso A: fecha_pago registrada en el futuro
    if fecha_p and fecha_p > hoy_str:
        if not tiene_ev:
            return True, "fecha_pago futura sin referencia (imposible)"
        return True, "fecha_pago futura con referencia (requiere revisión)"

    # Caso B: cuota futura marcada pagada sin evidencia (seed corrupto)
    if estado == "pagada" and fecha_c > hoy_str and not tiene_ev:
        return True, "cuota futura marcada pagada sin evidencia"

    return False, ""


# ─────────────────────── Función principal ────────────────────────────────────

def generar_excel(loanbooks: list[dict]) -> bytes:
    """Genera el archivo Excel con 2 hojas y retorna los bytes.

    Args:
        loanbooks: Lista de documentos loanbook (sin _id de MongoDB).

    Returns:
        bytes — contenido del archivo .xlsx listo para StreamingResponse.
    """
    hoy     = today_bogota()
    hoy_str = hoy.isoformat()

    wb = openpyxl.Workbook()

    # ═══════════════════════════════════════════
    # HOJA 1 — Créditos
    # ═══════════════════════════════════════════
    ws_cred = wb.active
    ws_cred.title = "Creditos"

    # Columnas en orden final especificado
    cols_cred = [
        "loanbook_id",           # 1  (idx 0)
        "Cliente",               # 2  (idx 1)
        "Cédula",                # 3  (idx 2)
        "Plan",                  # 4  (idx 3)
        "Modalidad pago",        # 5  (idx 4)
        "Cuota monto (DB)",      # 6  (idx 5)
        "Cuota inicial (DB)",    # 7  (idx 6)
        "Num cuotas (DB)",       # 8  (idx 7)  ← verde/rojo
        "Num cuotas (tabla)",    # 9  (idx 8)  ← verde/rojo
        "Cuotas OK",             # 10 (idx 9)  ← verde/rojo
        "Valor total (DB)",      # 11 (idx 10) ← verde/rojo
        "Valor total (tabla)",   # 12 (idx 11) ← verde/rojo
        "Total OK",              # 13 (idx 12) ← verde/rojo
        "Diferencia ($)",        # 14 (idx 13)
        "Estado",                # 15 (idx 14)
        "Fecha entrega",         # 16 (idx 15)
        "Total pagado",          # 17 (idx 16)
        "Saldo capital",         # 18 (idx 17)
        "Saldo intereses",       # 19 (idx 18)
        "Cartera total",         # 20 (idx 19) ← verde siempre
        "Capital plan",          # 21 (idx 20)
        "Cuota estándar",        # 22 (idx 21)
        "DPD",                   # 23 (idx 22) ← rojo si > 0
        "Mora acumulada ($)",    # 24 (idx 23) ← rojo si DPD > 0
        "Cuotas vencidas",       # 25 (idx 24)
        "Sub bucket",            # 26 (idx 25)
        "Teléfono",              # 27 (idx 26)
        "Ciudad",                # 28 (idx 27)
        "Fecha primer pago",     # 29 (idx 28)
        "Fecha vencimiento",     # 30 (idx 29)
        "Cuotas futuras pagadas",# 31 (idx 30) ← rojo si > 0
    ]
    _header_row(ws_cred, cols_cred)

    for lb in loanbooks:
        loanbook_id   = lb.get("loanbook_id", "?")
        cliente       = lb.get("cliente", {})
        nombre        = cliente.get("nombre", "?")
        cedula        = cliente.get("cedula", "?")
        telefono      = cliente.get("telefono") or ""
        ciudad        = cliente.get("ciudad") or ""
        plan_codigo   = lb.get("plan_codigo") or lb.get("plan", {}).get("codigo") or "?"
        modalidad     = lb.get("modalidad") or lb.get("plan", {}).get("modalidad") or "semanal"
        cuota_monto   = lb.get("cuota_monto") or lb.get("plan", {}).get("cuota_valor") or 0
        cuota_inicial = lb.get("plan", {}).get("cuota_inicial", 0) or 0
        num_cuotas_db = (
            lb.get("num_cuotas")
            or lb.get("plan", {}).get("total_cuotas")
            or len(lb.get("cuotas", []))
            or 0
        )
        valor_total_db  = lb.get("valor_total", 0) or 0
        estado          = lb.get("estado", "?")
        fecha_entrega   = lb.get("fecha_entrega", "")
        total_pagado    = lb.get("total_pagado", 0) or 0
        saldo_capital   = lb.get("saldo_capital", 0) or lb.get("saldo_pendiente", 0) or 0
        saldo_intereses = lb.get("saldo_intereses") or 0
        cartera_total   = (lb.get("saldo_capital") or 0) + (lb.get("saldo_intereses") or 0)
        capital_plan    = lb.get("capital_plan") or 0
        cuota_std       = lb.get("cuota_estandar_plan") or lb.get("cuota_monto") or 0
        dpd             = lb.get("dpd") or 0
        mora_cop        = lb.get("mora_acumulada_cop") or 0
        cuotas_venc     = lb.get("cuotas_vencidas") or 0
        sub_bucket      = lb.get("sub_bucket_semanal") or ""
        fecha_primer    = lb.get("fecha_primera_cuota") or lb.get("fecha_primer_pago") or ""
        fecha_venc      = lb.get("fecha_ultima_cuota") or lb.get("fecha_vencimiento") or ""

        # Valores correctos según tabla
        num_cuotas_ok_val  = get_num_cuotas(plan_codigo, modalidad)
        valor_total_ok_val = (
            get_valor_total(plan_codigo, modalidad, cuota_monto, cuota_inicial)
            if num_cuotas_ok_val is not None else None
        )

        cuotas_ok  = num_cuotas_ok_val is not None and num_cuotas_db == num_cuotas_ok_val
        total_ok   = valor_total_ok_val is not None and abs(valor_total_db - valor_total_ok_val) <= 1
        diferencia = (
            round(valor_total_db - valor_total_ok_val)
            if valor_total_ok_val is not None else "N/A"
        )

        # ¿Tiene cuotas futuras pagadas?
        cuotas_fut = sum(
            1 for c in lb.get("cuotas", [])
            if c.get("estado") == "pagada"
            and (c.get("fecha", "") or "") > hoy_str
            and not (bool(c.get("referencia")) or bool(c.get("metodo_pago") and c.get("metodo_pago") not in ("", "seed", None)))
        )

        row_idx = ws_cred.max_row + 1
        ws_cred.append([
            loanbook_id,                                        # 1
            nombre,                                             # 2
            cedula,                                             # 3
            plan_codigo,                                        # 4
            modalidad,                                          # 5
            cuota_monto,                                        # 6
            cuota_inicial,                                      # 7
            num_cuotas_db,                                      # 8
            num_cuotas_ok_val if num_cuotas_ok_val is not None else "N/A",  # 9
            "✓" if cuotas_ok else "✗",                         # 10
            valor_total_db,                                     # 11
            valor_total_ok_val if valor_total_ok_val is not None else "N/A",  # 12
            "✓" if total_ok else "✗",                          # 13
            diferencia,                                         # 14
            estado,                                             # 15
            fecha_entrega or "",                                # 16
            total_pagado,                                       # 17
            saldo_capital,                                      # 18
            saldo_intereses,                                    # 19
            cartera_total,                                      # 20
            capital_plan,                                       # 21
            cuota_std,                                          # 22
            dpd,                                                # 23
            mora_cop,                                           # 24
            cuotas_venc if cuotas_venc else "",                 # 25
            sub_bucket,                                         # 26
            telefono,                                           # 27
            ciudad,                                             # 28
            fecha_primer,                                       # 29
            fecha_venc,                                         # 30
            cuotas_fut if cuotas_fut > 0 else "",               # 31
        ])

        row = ws_cred[row_idx]

        # Num cuotas (idx 7-9)
        row[7].fill = _VERDE_FILL if cuotas_ok else _ROJO_FILL   # Num cuotas (DB)
        row[8].fill = _VERDE_FILL if cuotas_ok else _ROJO_FILL   # Num cuotas (tabla)
        row[9].fill = _VERDE_FILL if cuotas_ok else _ROJO_FILL   # Cuotas OK

        # Valor total (idx 10-12)
        row[10].fill = _VERDE_FILL if total_ok else _ROJO_FILL   # Valor total (DB)
        row[11].fill = _VERDE_FILL if total_ok else _ROJO_FILL   # Valor total (tabla)
        row[12].fill = _VERDE_FILL if total_ok else _ROJO_FILL   # Total OK

        # Cartera total (idx 19) — siempre verde para destacar
        row[19].fill = _VERDE_FILL

        # DPD y mora (idx 22-23) — rojo si hay mora
        if dpd and dpd > 0:
            row[22].fill = _ROJO_FILL   # DPD
            row[23].fill = _ROJO_FILL   # Mora acumulada ($)

        # Cuotas futuras pagadas (idx 30)
        if cuotas_fut > 0:
            row[30].fill = _ROJO_FILL

    _autofit(ws_cred)
    ws_cred.freeze_panes = "A2"

    # ═══════════════════════════════════════════
    # HOJA 2 — Cuotas (sin cambios)
    # ═══════════════════════════════════════════
    ws_cuotas = wb.create_sheet("Cuotas")

    cols_cuotas = [
        "loanbook_id",
        "Cliente",
        "# Cuota",
        "Monto",
        "Estado",
        "Fecha cuota",
        "Fecha pago",
        "Referencia",
        "Método pago",
        "Es corrupta",
        "Motivo corrupción",
    ]
    _header_row(ws_cuotas, cols_cuotas)

    for lb in loanbooks:
        loanbook_id = lb.get("loanbook_id", "?")
        nombre      = lb.get("cliente", {}).get("nombre", "?")
        cuotas      = lb.get("cuotas", [])

        for c in cuotas:
            es_corrupta, motivo = _cuota_corrupta(c, hoy_str)
            row_idx = ws_cuotas.max_row + 1
            ws_cuotas.append([
                loanbook_id,
                nombre,
                c.get("numero"),
                c.get("monto"),
                c.get("estado"),
                c.get("fecha", ""),
                c.get("fecha_pago", "") or "",
                c.get("referencia", "") or "",
                c.get("metodo_pago", "") or "",
                "✗ Sí" if es_corrupta else "✓ No",
                motivo,
            ])

            if es_corrupta:
                for cell in ws_cuotas[row_idx]:
                    cell.fill = _ROJO_FILL

    _autofit(ws_cuotas)
    ws_cuotas.freeze_panes = "A2"

    # Serializar a bytes
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
