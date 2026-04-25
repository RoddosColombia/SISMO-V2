"""
services/loanbook/loan_tape_service.py — Generador del Loan Tape en formato Excel.

Produce `loanbook_roddos_YYYY-MM-DD.xlsx` con 5 hojas:
  1. Loan Tape RDX     — Un loanbook RDX por fila, 38 columnas, celdas rojas en datos inválidos
  2. Loan Tape RODANTE — Un loanbook RODANTE por fila, columnas condicionales por subtipo
  3. Cronograma        — Una cuota por fila de todos los loanbooks, 20 columnas, color coding
  4. KPIs Mora         — 8 indicadores con valor, umbral y semáforo de color
  5. Roll Rate         — Matriz 5×5 de migración entre buckets (placeholder hasta historial)

Función pura: recibe loanbooks ya fetcheados, devuelve bytes del .xlsx.
Sin I/O de DB — el endpoint en routers/loanbook.py es responsable del fetch.

Ref: .planning/LOANBOOK_MAESTRO_v1.1.md caps 5, 10
"""

from __future__ import annotations

import io
from datetime import date, datetime, timedelta
from core.datetime_utils import now_bogota, today_bogota, now_iso_bogota
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ─────────────────────── Paleta de colores ────────────────────────────────────

_HDR_BG   = PatternFill("solid", fgColor="1F4E79")   # azul Roddos oscuro
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT     = Alignment(horizontal="left", vertical="center")

_FILL_ERROR   = PatternFill("solid", fgColor="FFCCCC")   # celda con dato inválido
_FILL_PAGADA  = PatternFill("solid", fgColor="90EE90")   # cuota pagada
_FILL_VENCIDA = PatternFill("solid", fgColor="FFB6C1")   # cuota vencida
_FILL_PROXIMA = PatternFill("solid", fgColor="FFFF99")   # cuota próxima
_FILL_DIAG    = PatternFill("solid", fgColor="DDEEFF")   # diagonal roll rate
_FILL_ALTO    = PatternFill("solid", fgColor="FFCCCC")   # celda > 5% roll rate
_FILL_CERO    = PatternFill("solid", fgColor="F5F5F5")   # cero

_COLOR_VERDE    = "00B050"
_COLOR_AMARILLO = "FFC000"
_COLOR_ROJO     = "FF0000"


# ─────────────────────── Helpers ──────────────────────────────────────────────

def _header_row(ws, columnas: list[str], row: int = 1) -> None:
    """Escribe fila de encabezados con estilo y ajusta anchos de columna."""
    for col_idx, nombre in enumerate(columnas, 1):
        cell = ws.cell(row=row, column=col_idx, value=nombre)
        cell.font = _HDR_FONT
        cell.fill = _HDR_BG
        cell.alignment = _CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(nombre) + 2)
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    ws.row_dimensions[row].height = 28


def _write_row(ws, row_idx: int, valores: list[Any]) -> None:
    """Escribe una fila de datos. Convierte dict/list a str para evitar errores."""
    for col_idx, val in enumerate(valores, 1):
        if isinstance(val, (dict, list)):
            val = str(val)
        ws.cell(row=row_idx, column=col_idx, value=val)


def _fill_row(ws, row_idx: int, n_cols: int, fill: PatternFill) -> None:
    """Aplica un fill a todas las celdas de una fila."""
    for col_idx in range(1, n_cols + 1):
        ws.cell(row=row_idx, column=col_idx).fill = fill


def _celda_roja(ws, row_idx: int, col_idx: int) -> None:
    ws.cell(row=row_idx, column=col_idx).fill = _FILL_ERROR


def _semaforo_color(valor: float, umbral: float, mayor_es_malo: bool = True) -> str:
    """Verde / amarillo / rojo según proximidad al umbral."""
    if mayor_es_malo:
        if valor <= umbral * 0.8:
            return _COLOR_VERDE
        if valor <= umbral:
            return _COLOR_AMARILLO
        return _COLOR_ROJO
    else:
        if valor >= umbral * 1.2:
            return _COLOR_VERDE
        if valor >= umbral:
            return _COLOR_AMARILLO
        return _COLOR_ROJO


def _v(doc: dict, *keys: str) -> Any:
    """Extrae valor anidado (soporta metadata_producto, cliente, plan)."""
    cur: Any = doc
    for k in keys:
        if isinstance(cur, dict):
            cur = cur.get(k)
        else:
            return None
    return cur


def _get(lb: dict, *paths: str, default: Any = None) -> Any:
    """Navega rutas anidadas con fallback dual.

    Ejemplo: _get(lb, 'fechas.factura', 'fecha_factura')
    Prueba cada path en orden (punto = subdocumento). Retorna el primero no-None.
    """
    for path in paths:
        val: Any = lb
        for key in path.split("."):
            if isinstance(val, dict):
                val = val.get(key)
            else:
                val = None
                break
        if val is not None:
            return val
    return default


def _proximo_miercoles(hoy: date) -> date:
    """Fecha del próximo miércoles (weekday=2) a partir de hoy."""
    dias_hasta = (2 - hoy.weekday()) % 7
    if dias_hasta == 0:
        dias_hasta = 7
    return hoy + timedelta(days=dias_hasta)


def _safe_float(val: Any, default: float = 0.0) -> float:
    try:
        return float(val) if val is not None else default
    except (TypeError, ValueError):
        return default


def _safe_int(val: Any, default: int = 0) -> int:
    try:
        return int(val) if val is not None else default
    except (TypeError, ValueError):
        return default


# ─────────────────────── Hoja 1: Loan Tape RDX ───────────────────────────────

_COLS_RDX = [
    # Identificación
    "loanbook_codigo", "producto", "cliente_nombre", "cliente_cedula",
    "cliente_telefono", "cliente_ciudad",
    # Términos
    "plan_codigo", "modalidad_pago", "fecha_factura", "fecha_entrega", "fecha_vencimiento",
    # Moto
    "moto_vin", "moto_modelo", "moto_motor", "moto_placa",
    "moto_anio", "moto_cilindraje", "moto_valor_origen", "ltv",
    # Montos
    "monto_original", "cuota_inicial", "cuota_periodica", "tasa_ea",
    # Desempeño
    "total_cuotas", "cuotas_pagadas", "cuotas_vencidas",
    "saldo_capital", "saldo_intereses", "mora_acumulada_cop", "dpd",
    # Estado
    "estado", "sub_bucket_semanal", "score_riesgo",
    # Enlaces
    "factura_alegra_id", "fecha_ultimo_pago", "vendedor", "whatsapp_status", "fecha_snapshot",
]

_IDX_RDX = {nombre: idx + 1 for idx, nombre in enumerate(_COLS_RDX)}


def _hoja_rdx(wb: openpyxl.Workbook, loanbooks: list[dict], fecha_corte: date) -> None:
    ws = wb.create_sheet("Loan Tape RDX")
    _header_row(ws, _COLS_RDX)

    for row_idx, lb in enumerate(loanbooks, 2):
        mp = lb.get("metadata_producto") or {}
        cliente = lb.get("cliente") or {}
        cuotas = lb.get("cuotas") or []
        pagadas = sum(1 for c in cuotas if c.get("estado") == "pagada")
        vencidas = sum(1 for c in cuotas if c.get("estado") in ("vencida", "parcial"))

        moto = lb.get("moto") or {}
        fila = [
            lb.get("loanbook_id") or lb.get("loanbook_codigo"),
            "RDX",
            cliente.get("nombre") or lb.get("cliente_nombre"),
            cliente.get("cedula") or lb.get("cliente_cedula"),
            cliente.get("telefono") or lb.get("cliente_telefono"),
            cliente.get("ciudad") or lb.get("cliente_ciudad"),
            _get(lb, "plan_codigo", "plan.codigo"),
            _get(lb, "modalidad_pago", "modalidad"),
            _get(lb, "fecha_factura", "fechas.factura"),
            _get(lb, "fecha_entrega", "fechas.entrega"),
            _get(lb, "fecha_vencimiento", "fechas.vencimiento"),
            # moto — fallback: metadata_producto → moto → top-level
            mp.get("moto_vin") or moto.get("vin") or lb.get("vin"),
            mp.get("moto_modelo") or moto.get("modelo") or lb.get("modelo"),
            mp.get("moto_motor") or moto.get("motor") or lb.get("motor"),
            mp.get("moto_placa") or moto.get("placa") or lb.get("placa"),
            mp.get("moto_anio") or moto.get("anio"),
            mp.get("moto_cilindraje") or moto.get("cilindraje"),
            mp.get("moto_valor_origen") or moto.get("valor_origen"),
            mp.get("ltv"),
            # montos
            _get(lb, "monto_original", "valor_total"),
            lb.get("cuota_inicial"),
            _get(lb, "cuota_periodica", "cuota_monto"),
            _get(lb, "tasa_ea", "plan.tasa", default=0.39),
            # desempeño
            _get(lb, "total_cuotas", "num_cuotas") or len(cuotas),
            pagadas,
            vencidas,
            lb.get("saldo_capital") or lb.get("saldo_pendiente"),
            lb.get("saldo_intereses", 0),
            lb.get("mora_acumulada_cop", 0),
            lb.get("dpd", 0),
            # estado
            lb.get("estado"),
            lb.get("sub_bucket_semanal"),
            lb.get("score_riesgo"),
            # enlaces
            lb.get("factura_alegra_id") or lb.get("alegra_factura_id"),
            lb.get("fecha_ultimo_pago"),
            lb.get("vendedor"),
            lb.get("whatsapp_status"),
            fecha_corte,
        ]
        _write_row(ws, row_idx, fila)

        # Celdas rojas por datos inválidos
        vin = mp.get("moto_vin") or lb.get("vin")
        if not vin:
            _celda_roja(ws, row_idx, _IDX_RDX["moto_vin"])

        saldo = _safe_float(lb.get("saldo_capital") or lb.get("saldo_pendiente"))
        if saldo < 0:
            _celda_roja(ws, row_idx, _IDX_RDX["saldo_capital"])

        dpd_val = _safe_int(lb.get("dpd", 0))
        estado_val = lb.get("estado", "")
        if dpd_val > 0 and estado_val == "Current":
            _celda_roja(ws, row_idx, _IDX_RDX["dpd"])
            _celda_roja(ws, row_idx, _IDX_RDX["estado"])


# ─────────────────────── Hoja 2: Loan Tape RODANTE ───────────────────────────

_COLS_RODANTE_BASE = [
    "loanbook_codigo", "producto", "subtipo_rodante",
    "cliente_nombre", "cliente_cedula", "cliente_telefono", "cliente_ciudad",
    "plan_codigo", "modalidad_pago", "fecha_factura", "fecha_entrega", "fecha_vencimiento",
    "monto_original", "cuota_inicial", "cuota_periodica", "tasa_ea",
    "total_cuotas", "cuotas_pagadas", "cuotas_vencidas",
    "saldo_capital", "saldo_intereses", "mora_acumulada_cop", "dpd",
    "estado", "sub_bucket_semanal", "score_riesgo",
    "factura_alegra_id", "fecha_ultimo_pago", "vendedor", "whatsapp_status", "fecha_snapshot",
    # repuestos
    "referencia_sku", "cantidad", "valor_unitario", "descripcion_repuesto",
    # soat
    "poliza_numero", "aseguradora", "vigencia_desde", "vigencia_hasta", "valor_soat", "placa_cubierta",
    # comparendo
    "comparendo_numero", "entidad_emisora", "fecha_infraccion", "valor_comparendo",
    # licencia
    "categoria_licencia", "centro_ensenanza_nombre", "fecha_inicio_curso", "valor_curso",
]


def _hoja_rodante(wb: openpyxl.Workbook, loanbooks: list[dict], fecha_corte: date) -> None:
    ws = wb.create_sheet("Loan Tape RODANTE")
    _header_row(ws, _COLS_RODANTE_BASE)

    for row_idx, lb in enumerate(loanbooks, 2):
        mp = lb.get("metadata_producto") or {}
        cliente = lb.get("cliente") or {}
        cuotas = lb.get("cuotas") or []
        pagadas = sum(1 for c in cuotas if c.get("estado") == "pagada")
        vencidas = sum(1 for c in cuotas if c.get("estado") in ("vencida", "parcial"))
        subtipo = lb.get("subtipo_rodante")

        fila = [
            lb.get("loanbook_id") or lb.get("loanbook_codigo"),
            "RODANTE",
            subtipo,
            cliente.get("nombre") or lb.get("cliente_nombre"),
            cliente.get("cedula") or lb.get("cliente_cedula"),
            cliente.get("telefono") or lb.get("cliente_telefono"),
            cliente.get("ciudad") or lb.get("cliente_ciudad"),
            _get(lb, "plan_codigo", "plan.codigo"),
            _get(lb, "modalidad_pago", "modalidad"),
            _get(lb, "fecha_factura", "fechas.factura"),
            _get(lb, "fecha_entrega", "fechas.entrega"),
            _get(lb, "fecha_vencimiento", "fechas.vencimiento"),
            _get(lb, "monto_original", "valor_total"),
            lb.get("cuota_inicial"),
            _get(lb, "cuota_periodica", "cuota_monto"),
            _get(lb, "tasa_ea", "plan.tasa", default=0.39),
            _get(lb, "total_cuotas", "num_cuotas") or len(cuotas),
            pagadas,
            vencidas,
            lb.get("saldo_capital") or lb.get("saldo_pendiente"),
            lb.get("saldo_intereses", 0),
            lb.get("mora_acumulada_cop", 0),
            lb.get("dpd", 0),
            lb.get("estado"),
            lb.get("sub_bucket_semanal"),
            lb.get("score_riesgo"),
            lb.get("factura_alegra_id") or lb.get("alegra_factura_id"),
            lb.get("fecha_ultimo_pago"),
            lb.get("vendedor"),
            lb.get("whatsapp_status"),
            fecha_corte,
            # repuestos
            mp.get("referencia_sku"),
            mp.get("cantidad"),
            mp.get("valor_unitario"),
            mp.get("descripcion_repuesto"),
            # soat
            mp.get("poliza_numero"),
            mp.get("aseguradora"),
            mp.get("vigencia_desde"),
            mp.get("vigencia_hasta"),
            mp.get("valor_soat"),
            mp.get("placa_cubierta"),
            # comparendo
            mp.get("comparendo_numero"),
            mp.get("entidad_emisora"),
            mp.get("fecha_infraccion"),
            mp.get("valor_comparendo"),
            # licencia
            mp.get("categoria_licencia"),
            mp.get("centro_ensenanza_nombre"),
            mp.get("fecha_inicio_curso"),
            mp.get("valor_curso"),
        ]
        _write_row(ws, row_idx, fila)


# ─────────────────────── Hoja 3: Cronograma ──────────────────────────────────

_COLS_CRONOGRAMA = [
    "loanbook_codigo", "cliente_nombre", "numero_cuota", "fecha_programada",
    "monto_total", "monto_capital", "monto_interes", "monto_fees", "estado",
    "fecha_pago", "monto_pagado", "metodo_pago", "banco", "referencia",
    "mora_acumulada", "mora_pagada", "anzi_pagado", "saldo_despues",
    "es_corrupta", "motivo_corrupcion",
]


def _hoja_cronograma(
    wb: openpyxl.Workbook,
    loanbooks: list[dict],
) -> None:
    ws = wb.create_sheet("Cronograma")
    _header_row(ws, _COLS_CRONOGRAMA)

    hoy = today_bogota()
    prox_miercoles = _proximo_miercoles(hoy)

    row_idx = 2
    for lb in loanbooks:
        lb_codigo = lb.get("loanbook_id") or lb.get("loanbook_codigo") or ""
        cliente = lb.get("cliente") or {}
        cliente_nombre = cliente.get("nombre") or lb.get("cliente_nombre") or ""
        cuotas = lb.get("cuotas") or []

        for cuota in cuotas:
            # Parsear fecha programada para comparar con próximo miércoles
            fecha_raw = (
                cuota.get("fecha_programada")
                or cuota.get("fecha")
            )
            fecha_programada = None
            if fecha_raw:
                if isinstance(fecha_raw, (date, datetime)):
                    fecha_programada = fecha_raw if isinstance(fecha_raw, date) else fecha_raw.date()
                else:
                    try:
                        fecha_programada = date.fromisoformat(str(fecha_raw)[:10])
                    except ValueError:
                        pass

            # Determinar si es corrupta
            monto_cuota = _safe_float(cuota.get("monto") or cuota.get("monto_total"))
            monto_pagado = _safe_float(cuota.get("monto_pagado", 0))
            estado_cuota = cuota.get("estado", "")
            es_corrupta = False
            motivo_corrupcion = ""
            if estado_cuota == "pagada" and monto_pagado <= 0:
                es_corrupta = True
                motivo_corrupcion = "Pagada sin monto_pagado"
            elif estado_cuota == "vencida" and not fecha_programada:
                es_corrupta = True
                motivo_corrupcion = "Vencida sin fecha_programada"

            fila = [
                lb_codigo,
                cliente_nombre,
                cuota.get("numero") or cuota.get("cuota_numero"),
                fecha_raw,
                monto_cuota,
                _safe_float(cuota.get("monto_capital", 0)),
                _safe_float(cuota.get("monto_interes", 0)),
                _safe_float(cuota.get("monto_fees", 0)),
                estado_cuota,
                cuota.get("fecha_pago"),
                monto_pagado,
                cuota.get("metodo_pago"),
                cuota.get("banco"),
                cuota.get("referencia"),
                _safe_float(cuota.get("mora_acumulada", 0) or cuota.get("mora_acumulada_cop", 0)),
                _safe_float(cuota.get("mora_pagada", 0)),
                _safe_float(cuota.get("anzi_pagado", 0)),
                cuota.get("saldo_despues"),
                "Sí" if es_corrupta else "No",
                motivo_corrupcion,
            ]
            _write_row(ws, row_idx, fila)

            # Color coding
            if estado_cuota == "pagada":
                _fill_row(ws, row_idx, len(_COLS_CRONOGRAMA), _FILL_PAGADA)
            elif estado_cuota == "vencida":
                _fill_row(ws, row_idx, len(_COLS_CRONOGRAMA), _FILL_VENCIDA)
            elif fecha_programada == prox_miercoles:
                _fill_row(ws, row_idx, len(_COLS_CRONOGRAMA), _FILL_PROXIMA)

            row_idx += 1


# ─────────────────────── Hoja 4: KPIs Mora ───────────────────────────────────

def _hoja_kpis_mora(wb: openpyxl.Workbook, loanbooks: list[dict]) -> None:
    ws = wb.create_sheet("KPIs Mora")
    _header_row(ws, ["Indicador", "Valor", "Umbral", "Estado"], row=1)

    activos = [lb for lb in loanbooks if lb.get("estado") not in ("Pagado", None)]
    total_n = len(activos) or 1
    total_saldo = sum(_safe_float(lb.get("saldo_capital") or lb.get("saldo_pendiente")) for lb in activos) or 1.0

    dpds = [_safe_int(lb.get("dpd", 0)) for lb in activos]
    mora_promedio = sum(dpds) / total_n

    activos_mora = [lb for lb in activos if _safe_int(lb.get("dpd", 0)) > 0]
    saldo_mora = sum(_safe_float(lb.get("saldo_capital") or lb.get("saldo_pendiente")) for lb in activos_mora)
    pct_mora = saldo_mora / total_saldo * 100

    intereses_mora = sum(_safe_int(lb.get("dpd", 0)) * 2_000 for lb in activos_mora)

    n_early = sum(1 for lb in activos if lb.get("estado") == "Early Delinquency")
    n_late_plus = sum(1 for lb in activos if lb.get("estado") in ("Late Delinquency", "Default", "Charge-Off"))
    n_default = sum(1 for lb in activos if lb.get("estado") == "Default")

    tasa_temprana = n_early / total_n * 100
    tasa_grave = n_late_plus / total_n * 100
    tasa_predefault = n_default / total_n * 100

    # Collection Rate: cuotas pagadas / cuotas debidas en el período
    total_cuotas_debidas = sum(
        sum(1 for c in (lb.get("cuotas") or []) if c.get("estado") in ("pagada", "vencida", "parcial"))
        for lb in loanbooks
    ) or 1
    total_cuotas_pagadas = sum(
        sum(1 for c in (lb.get("cuotas") or []) if c.get("estado") == "pagada")
        for lb in loanbooks
    )
    collection_rate = total_cuotas_pagadas / total_cuotas_debidas * 100

    kpis = [
        ("Días mora promedio",           mora_promedio,    5.0,         True),
        ("% Cartera en mora",            pct_mora,         15.0,        True),
        ("Valor en mora (COP)",          saldo_mora,       15_000_000,  True),
        ("Intereses mora pendientes (COP)", intereses_mora, 500_000,    True),
        ("Tasa mora temprana Early (%)", tasa_temprana,    20.0,        True),
        ("Tasa mora grave Late+ (%)",    tasa_grave,       10.0,        True),
        ("Tasa pre-default (%)",         tasa_predefault,  3.0,         True),
        ("Collection Rate (%)",          collection_rate,  95.0,        False),
    ]

    for row_idx, (nombre, valor, umbral, mayor_es_malo) in enumerate(kpis, 2):
        ws.cell(row=row_idx, column=1, value=nombre)
        ws.cell(row=row_idx, column=2, value=round(valor, 2))
        ws.cell(row=row_idx, column=3, value=umbral)
        color = _semaforo_color(valor, umbral, mayor_es_malo)
        semaforo_cell = ws.cell(row=row_idx, column=4, value="●")
        semaforo_cell.font = Font(color=color, size=16, bold=True)
        semaforo_cell.alignment = _CENTER

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 10


# ─────────────────────── Hoja 5: Roll Rate ───────────────────────────────────

_ROLL_ESTADOS = ["Current", "Early Delinq", "Mid Delinq", "Late Delinq", "Default/CO"]


def _hoja_roll_rate(wb: openpyxl.Workbook, loanbooks: list[dict]) -> None:
    ws = wb.create_sheet("Roll Rate")

    # Título
    title = ws.cell(row=1, column=1, value="Roll Rate — Migración semanal entre buckets")
    title.font = Font(bold=True, size=11)

    # Encabezados columna (estados destino)
    ws.cell(row=2, column=1, value="Semana anterior ↓ / actual →")
    ws.cell(row=2, column=1).font = Font(bold=True, italic=True, size=9)

    for col_idx, estado in enumerate(_ROLL_ESTADOS, 2):
        cell = ws.cell(row=2, column=col_idx, value=estado)
        cell.fill = _HDR_BG
        cell.font = _HDR_FONT
        cell.alignment = _CENTER

    # Encabezados fila (estados origen) + datos
    for row_idx, estado_origen in enumerate(_ROLL_ESTADOS, 3):
        row_label = ws.cell(row=row_idx, column=1, value=estado_origen)
        row_label.fill = _HDR_BG
        row_label.font = _HDR_FONT

        for col_idx in range(2, len(_ROLL_ESTADOS) + 2):
            # Placeholder 0.0 hasta que haya historial en loanbook_modificaciones
            val = 0.0
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.number_format = "0.0%"

            es_diagonal = (col_idx - 2) == (row_idx - 3)
            if es_diagonal:
                cell.fill = _FILL_DIAG
            elif val > 0.05:
                cell.fill = _FILL_ALTO
            else:
                cell.fill = _FILL_CERO
            cell.alignment = _CENTER

    # Nota sobre placeholder
    nota = ws.cell(
        row=len(_ROLL_ESTADOS) + 4,
        column=1,
        value=(
            "Nota: Roll Rate se calcula automáticamente cuando haya historial "
            "de transiciones en loanbook_modificaciones (disponible tras B2 en producción)."
        ),
    )
    nota.font = Font(italic=True, color="888888", size=9)

    ws.column_dimensions["A"].width = 22
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 14


# ─────────────────────── Función principal ───────────────────────────────────

def generar_loan_tape(
    loanbooks: list[dict],
    fecha_corte: Optional[date] = None,
) -> bytes:
    """Genera el Loan Tape Excel completo y retorna bytes del .xlsx.

    Función pura — sin I/O. El caller (endpoint HTTP) es responsable de
    fetchar los loanbooks desde MongoDB.

    Args:
        loanbooks:   lista de documentos de la colección loanbook (sin _id).
        fecha_corte: fecha del snapshot. Default: hoy.

    Returns:
        bytes del archivo .xlsx listo para enviar como HTTP response.
    """
    if fecha_corte is None:
        fecha_corte = today_bogota()

    rdx = [lb for lb in loanbooks if lb.get("producto") == "RDX"]
    rodante = [lb for lb in loanbooks if lb.get("producto") == "RODANTE"]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # eliminar hoja por defecto

    _hoja_rdx(wb, rdx, fecha_corte)
    _hoja_rodante(wb, rodante, fecha_corte)
    _hoja_cronograma(wb, loanbooks)
    _hoja_kpis_mora(wb, loanbooks)
    _hoja_roll_rate(wb, loanbooks)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
