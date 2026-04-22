"""
test_excel_export.py — Tests del export Excel (BUILD B).

2 tests:
  1. generar_excel() retorna bytes de un .xlsx válido (content-type check via magic bytes)
  2. Hoja "Creditos" tiene al menos 1 fila de datos; hoja "Cuotas" existe y tiene cabecera
"""
import io
import pytest
import openpyxl
from services.loanbook.excel_export import generar_excel


# ─── Fixtures ─────────────────────────────────────────────────────────────────

def _make_loanbook(loanbook_id="LB-TEST-01", plan_codigo="P39S", modalidad="semanal",
                   num_cuotas=39, cuota_monto=204_000, valor_total=None,
                   cuota_inicial=0, cuotas=None):
    return {
        "loanbook_id": loanbook_id,
        "cliente": {"nombre": "Test Cliente", "cedula": "12345678"},
        "plan_codigo": plan_codigo,
        "plan": {
            "codigo": plan_codigo,
            "modalidad": modalidad,
            "cuota_valor": cuota_monto,
            "total_cuotas": num_cuotas,
            "cuota_inicial": cuota_inicial,
        },
        "modalidad": modalidad,
        "cuota_monto": cuota_monto,
        "num_cuotas": num_cuotas,
        "valor_total": valor_total or (num_cuotas * cuota_monto + cuota_inicial),
        "cuotas": cuotas or [],
        "estado": "activo",
        "total_pagado": 0,
        "saldo_capital": num_cuotas * cuota_monto,
    }


# ─── Test 1: bytes válidos de xlsx ────────────────────────────────────────────

def test_generar_excel_produce_bytes_xlsx_validos():
    """generar_excel() retorna bytes que abren como workbook openpyxl válido."""
    lb = _make_loanbook()
    resultado = generar_excel([lb])

    assert isinstance(resultado, bytes), "Debe retornar bytes"
    assert len(resultado) > 0, "No debe estar vacío"

    # Magic bytes de ZIP (xlsx es un zip): PK\x03\x04
    assert resultado[:4] == b"PK\x03\x04", "Los bytes no tienen cabecera xlsx (PK\\x03\\x04)"

    # Debe abrirse como workbook sin error
    wb = openpyxl.load_workbook(io.BytesIO(resultado))
    assert "Creditos" in wb.sheetnames
    assert "Cuotas"   in wb.sheetnames


# ─── Test 2: hojas con datos correctos ───────────────────────────────────────

def test_generar_excel_hojas_con_datos_y_flags_corrupcion():
    """Hoja Creditos tiene fila por loanbook; Cuotas marca filas corruptas en rojo."""
    from datetime import date, timedelta

    hoy = date.today()
    manana = (hoy + timedelta(days=1)).isoformat()

    # LB limpio: P39S semanal, num_cuotas correcto
    lb_limpio = _make_loanbook("LB-LIMPIO", "P39S", "semanal", num_cuotas=39,
                               cuota_monto=204_000)

    # LB corrupto: num_cuotas incorrecto + cuota futura pagada sin referencia
    cuota_futura_pagada = {
        "numero": 35,
        "monto": 204_000,
        "estado": "pagada",
        "fecha": manana,
        "fecha_pago": manana,
        "referencia": None,
        "metodo_pago": None,
    }
    lb_corrupto = _make_loanbook(
        "LB-CORRUPTO", "P39S", "semanal",
        num_cuotas=28,         # incorrecto — debería ser 39
        cuota_monto=204_000,
        valor_total=28 * 204_000,  # incorrecto
        cuotas=[cuota_futura_pagada],
    )

    resultado = generar_excel([lb_limpio, lb_corrupto])
    wb = openpyxl.load_workbook(io.BytesIO(resultado))

    # ── Hoja Creditos ─────────────────────────────────────────────────────────
    ws_cred = wb["Creditos"]
    # Fila 1 = cabecera, filas 2+ = datos
    assert ws_cred.max_row >= 3, "Debe haber al menos 2 filas de datos (1 por loanbook)"

    # Primera columna de fila 2 debe ser loanbook_id
    ids_en_excel = [ws_cred.cell(row=r, column=1).value for r in range(2, ws_cred.max_row + 1)]
    assert "LB-LIMPIO"   in ids_en_excel
    assert "LB-CORRUPTO" in ids_en_excel

    # ── Hoja Cuotas ──────────────────────────────────────────────────────────
    ws_cuotas = wb["Cuotas"]
    assert ws_cuotas.max_row >= 2, "Debe haber al menos 1 fila de datos en Cuotas"

    # Buscar la cuota corrupta: columna "Es corrupta" debe ser "✗ Sí"
    col_es_corrupta = 10  # índice según cols_cuotas
    filas_corruptas = [
        ws_cuotas.cell(row=r, column=col_es_corrupta).value
        for r in range(2, ws_cuotas.max_row + 1)
        if ws_cuotas.cell(row=r, column=col_es_corrupta).value == "✗ Sí"
    ]
    assert len(filas_corruptas) >= 1, "Debe haber al menos 1 cuota marcada como corrupta"
