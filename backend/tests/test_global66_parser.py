"""
Tests for Global66 bank parser.
Uses openpyxl to create temporary .xlsx files with the expected structure.
"""
import pytest
import openpyxl
from services.bank_parsers import parse_global66, detect_bank


def _create_global66_xlsx(path, data_rows):
    """Helper: create a minimal Global66-format xlsx at the given path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos de cuenta COP"

    # Row 1: title
    ws.cell(row=1, column=1, value="Movimientos de cuenta COP")

    # Row 2: period
    ws.cell(row=2, column=1, value="Periodo: 2026-03-01 a 2026-03-31")

    # Row 4: headers (14 columns)
    headers = [
        "Tipo transaccion", "Fecha", "Monto debitado", "Monto acreditado",
        "Col E", "Col F", "Col G", "Nombre tercero",
        "DNI tercero", "Col J", "Col K", "Col L",
        "ID transaccion", "Comentario",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col_idx, value=h)

    # Data rows starting at row 5
    for row_offset, row_data in enumerate(data_rows):
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=5 + row_offset, column=col_idx, value=val)

    wb.save(str(path))
    wb.close()


class TestParseGlobal66Egreso:
    def test_parse_global66_egreso(self, tmp_path):
        """Debit row: monto in col C (index 2), col D empty."""
        xlsx_path = tmp_path / "global66_egreso.xlsx"
        # 14 columns: A..N
        row = [
            "Debito",                    # A: Tipo transaccion
            "2026-03-15 10:30:00",       # B: Fecha
            150000.0,                    # C: Monto debitado
            None,                        # D: Monto acreditado
            None, None, None,            # E, F, G
            "RESTAURANTE XYZ",           # H: Nombre tercero
            "900123456",                 # I: DNI tercero
            None, None, None,            # J, K, L
            "TXN-001-ABC",              # M: ID transaccion
            "PAGO ALMUERZO EQUIPO",      # N: Comentario
        ]
        _create_global66_xlsx(xlsx_path, [row])

        movements = parse_global66(str(xlsx_path))

        assert len(movements) == 1
        m = movements[0]
        assert m["fecha"] == "2026-03-15"
        assert m["tipo"] == "debito"
        assert m["monto"] == 150000.0
        assert m["banco"] == "Global66"
        assert m["referencia"] == "TXN-001-ABC"
        assert "Debito" in m["descripcion"]
        assert "PAGO ALMUERZO EQUIPO" in m["descripcion"]
        assert "RESTAURANTE XYZ" in m["descripcion"]


class TestParseGlobal66Ingreso:
    def test_parse_global66_ingreso(self, tmp_path):
        """Credit row: monto in col D (index 3), col C empty."""
        xlsx_path = tmp_path / "global66_ingreso.xlsx"
        row = [
            "Abono",                     # A: Tipo transaccion
            "2026-03-20 14:00:00",       # B: Fecha
            None,                        # C: Monto debitado (empty)
            2500000.0,                   # D: Monto acreditado
            None, None, None,            # E, F, G
            "CLIENTE JUAN PEREZ",        # H: Nombre tercero
            "1020304050",                # I: DNI tercero
            None, None, None,            # J, K, L
            "TXN-002-DEF",              # M: ID transaccion
            "PAGO FACTURA #123",         # N: Comentario
        ]
        _create_global66_xlsx(xlsx_path, [row])

        movements = parse_global66(str(xlsx_path))

        assert len(movements) == 1
        m = movements[0]
        assert m["fecha"] == "2026-03-20"
        assert m["tipo"] == "credito"
        assert m["monto"] == 2500000.0
        assert m["banco"] == "Global66"
        assert m["referencia"] == "TXN-002-DEF"
        assert "Abono" in m["descripcion"]
        assert "PAGO FACTURA #123" in m["descripcion"]
        assert "CLIENTE JUAN PEREZ" in m["descripcion"]


class TestDetectBankGlobal66:
    def test_detect_bank_global66(self, tmp_path):
        """detect_bank returns 'global66' for xlsx with sheet 'Movimientos de cuenta COP'."""
        xlsx_path = tmp_path / "global66_detect.xlsx"
        _create_global66_xlsx(xlsx_path, [])

        result = detect_bank(str(xlsx_path))
        assert result == "global66"


class TestParseGlobal66SkipsEmptyRows:
    def test_skips_rows_without_values(self, tmp_path):
        """Rows where both C and D are empty/None should be skipped."""
        xlsx_path = tmp_path / "global66_empty.xlsx"
        row_empty = [
            "GMF (4x1.000)", "2026-03-10 08:00:00",
            None, None,  # Both debit and credit empty
            None, None, None, "BANCO", "123", None, None, None, "TXN-003", "GMF",
        ]
        row_valid = [
            "Debito", "2026-03-10 09:00:00",
            5000.0, None,
            None, None, None, "TIENDA", "456", None, None, None, "TXN-004", "COMPRA",
        ]
        _create_global66_xlsx(xlsx_path, [row_empty, row_valid])

        movements = parse_global66(str(xlsx_path))

        assert len(movements) == 1
        assert movements[0]["referencia"] == "TXN-004"
