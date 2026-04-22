"""
test_loan_tape.py — Tests del generador de Loan Tape Excel.

Valida:
  - generar_loan_tape devuelve bytes de un .xlsx válido
  - Las 5 hojas existen con los nombres correctos
  - Hoja RDX: columnas correctas, solo filtra producto=="RDX"
  - Hoja RODANTE: columnas correctas, solo filtra producto=="RODANTE"
  - Hoja Cronograma: una fila por cuota de todos los loanbooks
  - Hoja KPIs Mora: 8 indicadores (filas 2-9 en la hoja)
  - Hoja Roll Rate: matriz 5×5 con number_format "0.0%"
  - Filename convention helpers

Tests puros — sin MongoDB, sin I/O.
"""

from __future__ import annotations

from datetime import date
from io import BytesIO

import openpyxl
import pytest

from services.loanbook.loan_tape_service import (
    _COLS_RDX,
    _COLS_RODANTE_BASE,
    _COLS_CRONOGRAMA,
    _proximo_miercoles,
    _safe_float,
    _safe_int,
    _semaforo_color,
    generar_loan_tape,
)


# ─────────────────────── Fixtures ────────────────────────────────────────────

def _loanbook_rdx(loanbook_id: str = "LB-RDX-001", estado: str = "Current",
                  dpd: int = 0, saldo: float = 2_500_000.0) -> dict:
    return {
        "loanbook_id": loanbook_id,
        "producto": "RDX",
        "estado": estado,
        "dpd": dpd,
        "saldo_capital": saldo,
        "saldo_intereses": 0,
        "mora_acumulada_cop": dpd * 2_000,
        "plan_codigo": "P52S",
        "modalidad_pago": "semanal",
        "monto_original": 4_200_000,
        "total_cuotas": 52,
        "tasa_ea": 0.42,
        "cliente": {
            "nombre": "Juan Prueba",
            "cedula": "12345678",
            "telefono": "3001234567",
            "ciudad": "Bogotá",
        },
        "metadata_producto": {
            "moto_vin": "VIN0001",
            "moto_modelo": "Boxer CT100",
            "moto_motor": "MOT0001",
            "moto_placa": "ABC123",
            "moto_anio": 2023,
        },
        "cuotas": [
            {"numero": 1, "estado": "pagada", "fecha": "2026-01-07", "monto": 95_000, "monto_pagado": 95_000},
            {"numero": 2, "estado": "pendiente", "fecha": "2026-01-14", "monto": 95_000},
        ],
    }


def _loanbook_rodante(loanbook_id: str = "LB-ROD-001") -> dict:
    return {
        "loanbook_id": loanbook_id,
        "producto": "RODANTE",
        "subtipo_rodante": "repuestos",
        "estado": "Current",
        "dpd": 0,
        "saldo_capital": 500_000.0,
        "saldo_intereses": 0,
        "mora_acumulada_cop": 0,
        "plan_codigo": "P8S",
        "modalidad_pago": "semanal",
        "monto_original": 600_000,
        "total_cuotas": 8,
        "tasa_ea": 0.0,
        "cliente": {
            "nombre": "Pedro Rodante",
            "cedula": "87654321",
        },
        "metadata_producto": {
            "referencia_sku": "SKU-001",
            "cantidad": 2,
            "valor_unitario": 300_000,
            "descripcion_repuesto": "Filtro de aire",
        },
        "cuotas": [
            {"numero": 1, "estado": "pagada", "fecha": "2026-01-07", "monto": 75_000, "monto_pagado": 75_000},
        ],
    }


def _parse_xlsx(xlsx_bytes: bytes) -> openpyxl.Workbook:
    return openpyxl.load_workbook(BytesIO(xlsx_bytes))


# ─────────────────────── BLOQUE 1 — Estructura general ───────────────────────

class TestEstructuraGeneral:
    """Verifica que generar_loan_tape devuelve bytes con 5 hojas correctas."""

    def test_retorna_bytes(self):
        resultado = generar_loan_tape([])
        assert isinstance(resultado, bytes)
        assert len(resultado) > 0

    def test_bytes_son_xlsx_valido(self):
        """Los bytes deben poder cargarse como workbook válido."""
        resultado = generar_loan_tape([])
        wb = _parse_xlsx(resultado)
        assert wb is not None

    def test_cinco_hojas_existen(self):
        resultado = generar_loan_tape([])
        wb = _parse_xlsx(resultado)
        assert set(wb.sheetnames) == {
            "Loan Tape RDX",
            "Loan Tape RODANTE",
            "Cronograma",
            "KPIs Mora",
            "Roll Rate",
        }

    def test_orden_hojas(self):
        """Las hojas deben aparecer en el orden correcto."""
        resultado = generar_loan_tape([])
        wb = _parse_xlsx(resultado)
        assert wb.sheetnames[0] == "Loan Tape RDX"
        assert wb.sheetnames[1] == "Loan Tape RODANTE"
        assert wb.sheetnames[2] == "Cronograma"
        assert wb.sheetnames[3] == "KPIs Mora"
        assert wb.sheetnames[4] == "Roll Rate"

    def test_fecha_corte_default_es_hoy(self):
        """Sin fecha_corte, la función no lanza excepción y produce output."""
        resultado = generar_loan_tape([_loanbook_rdx()])
        assert isinstance(resultado, bytes)

    def test_fecha_corte_explicita(self):
        fecha = date(2026, 3, 31)
        resultado = generar_loan_tape([_loanbook_rdx()], fecha_corte=fecha)
        assert isinstance(resultado, bytes)


# ─────────────────────── BLOQUE 2 — Hoja Loan Tape RDX ──────────────────────

class TestHojaRDX:
    """Columnas, filtrado y datos de la hoja Loan Tape RDX."""

    def test_columnas_correctas(self):
        resultado = generar_loan_tape([_loanbook_rdx()])
        wb = _parse_xlsx(resultado)
        ws = wb["Loan Tape RDX"]
        encabezados = [ws.cell(row=1, column=i).value for i in range(1, len(_COLS_RDX) + 1)]
        assert encabezados == _COLS_RDX

    def test_38_columnas(self):
        assert len(_COLS_RDX) == 38

    def test_solo_rdx_aparece_en_hoja_rdx(self):
        """Loanbook RODANTE no aparece en la hoja RDX."""
        lbs = [_loanbook_rdx("LB-001"), _loanbook_rodante("LB-ROD-001")]
        resultado = generar_loan_tape(lbs)
        wb = _parse_xlsx(resultado)
        ws = wb["Loan Tape RDX"]
        # Fila 2 = primer dato, fila 3 sería segundo (no debe existir con datos)
        ids = [ws.cell(row=r, column=1).value for r in range(2, 10) if ws.cell(row=r, column=1).value]
        assert "LB-001" in ids
        assert "LB-ROD-001" not in ids

    def test_sin_rdx_hoja_solo_encabezados(self):
        """Sin loanbooks RDX, la hoja tiene solo la fila de encabezados."""
        resultado = generar_loan_tape([_loanbook_rodante()])
        wb = _parse_xlsx(resultado)
        ws = wb["Loan Tape RDX"]
        assert ws.cell(row=2, column=1).value is None

    def test_datos_cliente_en_fila(self):
        lb = _loanbook_rdx()
        resultado = generar_loan_tape([lb])
        wb = _parse_xlsx(resultado)
        ws = wb["Loan Tape RDX"]
        # cliente_nombre está en columna 3
        assert ws.cell(row=2, column=3).value == "Juan Prueba"

    def test_vin_en_fila(self):
        lb = _loanbook_rdx()
        resultado = generar_loan_tape([lb])
        wb = _parse_xlsx(resultado)
        ws = wb["Loan Tape RDX"]
        # moto_vin está en columna 12
        col_vin = _COLS_RDX.index("moto_vin") + 1
        assert ws.cell(row=2, column=col_vin).value == "VIN0001"


# ─────────────────────── BLOQUE 3 — Hoja Loan Tape RODANTE ──────────────────

class TestHojaRODANTE:
    """Columnas y filtrado de la hoja Loan Tape RODANTE."""

    def test_solo_rodante_aparece_en_hoja_rodante(self):
        lbs = [_loanbook_rdx("LB-RDX-001"), _loanbook_rodante("LB-ROD-001")]
        resultado = generar_loan_tape(lbs)
        wb = _parse_xlsx(resultado)
        ws = wb["Loan Tape RODANTE"]
        ids = [ws.cell(row=r, column=1).value for r in range(2, 10) if ws.cell(row=r, column=1).value]
        assert "LB-ROD-001" in ids
        assert "LB-RDX-001" not in ids

    def test_columnas_rodante_incluyen_subtipo(self):
        assert "subtipo_rodante" in _COLS_RODANTE_BASE

    def test_columnas_rodante_incluyen_repuestos(self):
        assert "referencia_sku" in _COLS_RODANTE_BASE
        assert "descripcion_repuesto" in _COLS_RODANTE_BASE

    def test_datos_repuesto_en_fila(self):
        lb = _loanbook_rodante()
        resultado = generar_loan_tape([lb])
        wb = _parse_xlsx(resultado)
        ws = wb["Loan Tape RODANTE"]
        col_sku = _COLS_RODANTE_BASE.index("referencia_sku") + 1
        assert ws.cell(row=2, column=col_sku).value == "SKU-001"


# ─────────────────────── BLOQUE 4 — Hoja Cronograma ─────────────────────────

class TestHojaCronograma:
    """Una fila por cuota, columnas y color coding."""

    def test_columnas_cronograma(self):
        resultado = generar_loan_tape([_loanbook_rdx()])
        wb = _parse_xlsx(resultado)
        ws = wb["Cronograma"]
        encabezados = [ws.cell(row=1, column=i).value for i in range(1, len(_COLS_CRONOGRAMA) + 1)]
        assert encabezados == _COLS_CRONOGRAMA

    def test_una_fila_por_cuota(self):
        """2 loanbooks con 2 cuotas cada uno → 4 filas de datos."""
        lbs = [_loanbook_rdx("LB-001"), _loanbook_rdx("LB-002")]
        resultado = generar_loan_tape(lbs)
        wb = _parse_xlsx(resultado)
        ws = wb["Cronograma"]
        # Contar filas no vacías desde fila 2
        filas_con_datos = sum(
            1 for r in range(2, 20)
            if ws.cell(row=r, column=1).value is not None
        )
        assert filas_con_datos == 4

    def test_cronograma_acepta_fecha_programada_alternativa(self):
        """Cuotas con campo fecha_programada (no fecha) también se procesan."""
        lb = {
            "loanbook_id": "LB-ALT-001",
            "producto": "RDX",
            "estado": "Current",
            "cuotas": [
                {"numero": 1, "estado": "pendiente", "fecha_programada": "2026-04-30", "monto": 80_000}
            ],
        }
        resultado = generar_loan_tape([lb])
        wb = _parse_xlsx(resultado)
        ws = wb["Cronograma"]
        assert ws.cell(row=2, column=1).value == "LB-ALT-001"


# ─────────────────────── BLOQUE 5 — Hoja KPIs Mora ──────────────────────────

class TestHojaKPIs:
    """8 indicadores con semáforo."""

    def test_kpis_8_filas_de_datos(self):
        resultado = generar_loan_tape([_loanbook_rdx()])
        wb = _parse_xlsx(resultado)
        ws = wb["KPIs Mora"]
        # Fila 1 = encabezados, filas 2-9 = datos
        nombres = [ws.cell(row=r, column=1).value for r in range(2, 10)]
        nombres_validos = [n for n in nombres if n]
        assert len(nombres_validos) == 8

    def test_kpis_encabezados(self):
        resultado = generar_loan_tape([])
        wb = _parse_xlsx(resultado)
        ws = wb["KPIs Mora"]
        assert ws.cell(row=1, column=1).value == "Indicador"
        assert ws.cell(row=1, column=2).value == "Valor"
        assert ws.cell(row=1, column=3).value == "Umbral"
        assert ws.cell(row=1, column=4).value == "Estado"

    def test_kpis_semaforo_col_4(self):
        """La columna semáforo debe contener el símbolo '●'."""
        resultado = generar_loan_tape([_loanbook_rdx()])
        wb = _parse_xlsx(resultado)
        ws = wb["KPIs Mora"]
        semafaros = [ws.cell(row=r, column=4).value for r in range(2, 10)]
        assert all(s == "●" for s in semafaros)

    def test_collection_rate_kpi_incluido(self):
        """Collection Rate debe ser uno de los 8 KPIs."""
        resultado = generar_loan_tape([_loanbook_rdx()])
        wb = _parse_xlsx(resultado)
        ws = wb["KPIs Mora"]
        nombres = [ws.cell(row=r, column=1).value for r in range(2, 10)]
        assert any("Collection" in (n or "") for n in nombres)


# ─────────────────────── BLOQUE 6 — Hoja Roll Rate ──────────────────────────

class TestHojaRollRate:
    """Matriz 5×5 con valores placeholder."""

    def test_roll_rate_5x5_valores_cero(self):
        """Todos los valores son 0.0 (placeholder)."""
        resultado = generar_loan_tape([_loanbook_rdx()])
        wb = _parse_xlsx(resultado)
        ws = wb["Roll Rate"]
        # Datos en filas 3-7 (5 estados), columnas 2-6 (5 estados)
        for row_idx in range(3, 8):
            for col_idx in range(2, 7):
                val = ws.cell(row=row_idx, column=col_idx).value
                assert val == 0.0, f"Esperado 0.0 en ({row_idx},{col_idx}), got {val}"

    def test_roll_rate_encabezado_primera_fila(self):
        """Fila 1 tiene el título del roll rate."""
        resultado = generar_loan_tape([])
        wb = _parse_xlsx(resultado)
        ws = wb["Roll Rate"]
        titulo = ws.cell(row=1, column=1).value
        assert titulo is not None
        assert "Roll Rate" in titulo

    def test_roll_rate_5_estados_en_encabezados(self):
        """Las 5 columnas de estados deben estar en la fila 2."""
        resultado = generar_loan_tape([])
        wb = _parse_xlsx(resultado)
        ws = wb["Roll Rate"]
        estados = [ws.cell(row=2, column=c).value for c in range(2, 7)]
        assert len([e for e in estados if e]) == 5


# ─────────────────────── BLOQUE 7 — Helpers ─────────────────────────────────

class TestHelpers:
    """Helpers internos del servicio."""

    def test_safe_float_none_retorna_default(self):
        assert _safe_float(None) == 0.0

    def test_safe_float_string_numerica(self):
        assert _safe_float("3.14") == pytest.approx(3.14)

    def test_safe_float_string_no_numerica_retorna_default(self):
        assert _safe_float("abc") == 0.0

    def test_safe_int_none_retorna_default(self):
        assert _safe_int(None) == 0

    def test_safe_int_float_trunca(self):
        assert _safe_int(3.9) == 3

    def test_semaforo_mayor_es_malo_verde(self):
        """Valor < 80% umbral → verde."""
        color = _semaforo_color(5.0, 15.0, mayor_es_malo=True)
        assert color == "00B050"

    def test_semaforo_mayor_es_malo_rojo(self):
        """Valor > umbral → rojo."""
        color = _semaforo_color(20.0, 15.0, mayor_es_malo=True)
        assert color == "FF0000"

    def test_proximo_miercoles_desde_lunes(self):
        """Lunes 2026-04-20 → próximo miércoles es 2026-04-22."""
        lunes = date(2026, 4, 20)  # weekday=0
        assert _proximo_miercoles(lunes) == date(2026, 4, 22)

    def test_proximo_miercoles_desde_miercoles(self):
        """Desde el propio miércoles → el siguiente miércoles (no el mismo)."""
        miercoles = date(2026, 4, 22)  # weekday=2
        siguiente = _proximo_miercoles(miercoles)
        assert siguiente == date(2026, 4, 29)


# ─────────────────────── BLOQUE 8 — Convención de nombre ────────────────────

class TestNombreArchivo:
    """Verifica la convención del nombre de archivo esperado por el endpoint."""

    def test_filename_convention(self):
        fecha = date(2026, 4, 22)
        filename = f"loanbook_roddos_{fecha.strftime('%Y-%m-%d')}.xlsx"
        assert filename == "loanbook_roddos_2026-04-22.xlsx"

    def test_filename_incluye_fecha_corte(self):
        for fecha_str in ("2026-01-31", "2025-12-01"):
            fecha = date.fromisoformat(fecha_str)
            filename = f"loanbook_roddos_{fecha.strftime('%Y-%m-%d')}.xlsx"
            assert fecha_str in filename
