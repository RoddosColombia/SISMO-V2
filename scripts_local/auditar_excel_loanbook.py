"""
Auditoria COMPLETA del Excel oficial RODDOS Loanbook.

Lee el Excel sin asumir nada y reporta:
  1. Hojas presentes y sus dimensiones.
  2. Headers de cada hoja con su letra de columna (A, B, C ... AA, AB).
  3. Por cada hoja con LBs:
       - cuota_inicial (col U esperada): cuántos > 0, cuántos = 0, suma
       - monto_original (col T): suma
       - saldo_capital (col AA): suma
       - saldo_intereses (col AB): suma
       - Validación: T = AA + AB (capital + intereses)?
  4. Detalle por LB: id, nombre, cuota_inicial, monto_original, saldo_capital, saldo_intereses
  5. Detección de fórmulas vs valores (data_only=False y data_only=True)
  6. Genera .planning/marcar_cuotas_iniciales_2026-04-30.json con los LBs CI>0.

Uso:
    python scripts_local/auditar_excel_loanbook.py [<ruta_xlsx>]

Si no se da ruta usa loanbook_roddos_2026-04-30.xlsx en cwd.
"""
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


def col_letter(idx_1based: int) -> str:
    """Convierte índice 1-based a letra de columna Excel (1='A', 27='AA')."""
    return get_column_letter(idx_1based)


def fecha_iso(v):
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    return str(v) if v else None


def auditar_hoja(ws, esperados=("cuota_inicial", "monto_original",
                                 "saldo_capital", "saldo_intereses",
                                 "loanbook_codigo", "cliente_nombre",
                                 "fecha_entrega", "estado")):
    print(f"\n========== HOJA: {ws.title} ({ws.max_row} filas x {ws.max_column} cols) ==========")

    # Headers con letra de columna
    headers = []
    for i in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=i).value
        headers.append(v)
        letter = col_letter(i)
        marker = " <--" if v in esperados else ""
        print(f"  {letter:>3}  col{i:<3} = {v!r}{marker}")

    # Resolver índices de las columnas esperadas
    idx = {}
    for k in esperados:
        try:
            idx[k] = headers.index(k)
        except ValueError:
            idx[k] = None

    return headers, idx


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("loanbook_roddos_2026-04-30.xlsx")
    if not xlsx.exists():
        print(f"ERROR: {xlsx} no existe.")
        print(f"  cwd: {Path.cwd()}")
        sys.exit(1)

    print(f"Archivo: {xlsx.resolve()}")
    print(f"Tamano:  {xlsx.stat().st_size:,} bytes")

    # 1) Cargar dos veces: con valores calculados y con formulas
    wb_vals = openpyxl.load_workbook(xlsx, data_only=True)
    wb_fmt  = openpyxl.load_workbook(xlsx, data_only=False)

    print(f"\nHojas: {wb_vals.sheetnames}")

    # 2) Detectar fórmulas en hojas con LBs
    print("\n========== Detección de fórmulas ==========")
    for sn in wb_fmt.sheetnames:
        ws = wb_fmt[sn]
        formulas = 0
        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 50)):
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    formulas += 1
        print(f"  {sn}: {formulas} fórmulas detectadas (en primeras 50 filas)")

    # 3) Auditoría por hoja
    records_ci = []
    total_ci = 0
    total_mo = 0
    total_sc = 0
    total_si = 0
    sin_ci   = []

    for sn in wb_vals.sheetnames:
        if sn not in ("Loan Tape RDX", "Loan Tape RODANTE"):
            print(f"\n========== {sn} ==========")
            ws = wb_vals[sn]
            print(f"  {ws.max_row} filas x {ws.max_column} cols (skipping audit)")
            continue

        ws = wb_vals[sn]
        headers, idx = auditar_hoja(ws)

        ci_col = idx.get("cuota_inicial")
        mo_col = idx.get("monto_original")
        sc_col = idx.get("saldo_capital")
        si_col = idx.get("saldo_intereses")
        id_col = idx.get("loanbook_codigo")
        nom_col = idx.get("cliente_nombre")
        fent_col = idx.get("fecha_entrega")
        est_col = idx.get("estado")

        if id_col is None:
            print(f"  ATENCION: hoja sin loanbook_codigo, skip")
            continue

        sub_total_ci = 0
        sub_total_mo = 0
        sub_total_sc = 0
        sub_total_si = 0
        sub_lbs_ci = 0
        sub_lbs_no_ci = 0

        print(f"\n  Detalle por LB:")
        print(f"  {'LB':<14} {'cliente':<32} {'CI(U)':>13} {'MO(T)':>13} {'SC(AA)':>13} {'SI(AB)':>13} {'T=AA+AB?':<10} fent")

        for row in ws.iter_rows(min_row=2, values_only=True):
            if id_col is None or not row[id_col]:
                continue
            lb_id = row[id_col]
            cliente = row[nom_col] if nom_col is not None else "?"

            ci = int(row[ci_col] or 0) if ci_col is not None else 0
            mo = int(row[mo_col] or 0) if mo_col is not None else 0
            sc = int(row[sc_col] or 0) if sc_col is not None else 0
            si = int(row[si_col] or 0) if si_col is not None else 0
            fent = row[fent_col] if fent_col is not None else None
            est  = row[est_col] if est_col is not None else None

            sub_total_mo += mo
            sub_total_sc += sc
            sub_total_si += si

            check = "OK" if (sc + si) == mo else f"diff={mo-(sc+si):+,}"

            if ci > 0:
                sub_total_ci += ci
                sub_lbs_ci += 1
                records_ci.append({
                    "loanbook_id":   lb_id,
                    "cliente":       cliente,
                    "cuota_inicial": ci,
                    "fecha_pago":    fecha_iso(fent),
                    "metodo_pago":   "cuota_inicial_pre_entrega",
                    "sheet":         sn,
                    "estado_excel":  est,
                    "monto_original": mo,
                    "saldo_capital":  sc,
                    "saldo_intereses": si,
                })
            else:
                sub_lbs_no_ci += 1
                sin_ci.append({"lb": lb_id, "cliente": cliente, "estado": est,
                               "mo": mo, "sc": sc, "si": si, "fent": fecha_iso(fent)})

            print(f"  {str(lb_id):<14} {(str(cliente) or '?')[:30]:<32} "
                  f"${ci:>12,} ${mo:>12,} ${sc:>12,} ${si:>12,} {check:<10} {fecha_iso(fent)}")

        print(f"\n  --- Subtotales {sn} ---")
        print(f"  LBs con CI>0:        {sub_lbs_ci}")
        print(f"  LBs con CI=0:        {sub_lbs_no_ci}")
        print(f"  Sigma cuota_inicial: ${sub_total_ci:,}")
        print(f"  Sigma monto_original:${sub_total_mo:,}")
        print(f"  Sigma saldo_capital: ${sub_total_sc:,}")
        print(f"  Sigma saldo_int:     ${sub_total_si:,}")
        print(f"  saldo_cap + int:     ${sub_total_sc + sub_total_si:,}  (vs MO: diff {sub_total_mo - (sub_total_sc + sub_total_si):+,})")

        total_ci += sub_total_ci
        total_mo += sub_total_mo
        total_sc += sub_total_sc
        total_si += sub_total_si

    # 4) Totales globales
    print("\n========== TOTALES GLOBALES ==========")
    print(f"  Sigma cuota_inicial (col U):    ${total_ci:,}")
    print(f"  Sigma monto_original (col T):   ${total_mo:,}")
    print(f"  Sigma saldo_capital (col AA):   ${total_sc:,}")
    print(f"  Sigma saldo_intereses (col AB): ${total_si:,}")
    print(f"  saldo_cap + intereses:          ${total_sc + total_si:,}")
    print()
    print(f"  Andres dice:")
    print(f"    Sigma cuota_inicial:          $47,830,000  ({34} LBs)")
    print(f"    Sigma capital + intereses:    $368,774,281")
    print(f"    Total cartera (CI+cap+int):   $416,604,281")
    print()

    coincide_ci   = total_ci == 47_830_000
    coincide_mo   = total_mo == 368_774_281
    coincide_total = (total_ci + total_mo) == 416_604_281

    print(f"  Coincide CI:                    {coincide_ci}")
    print(f"  Coincide capital+intereses:     {coincide_mo}")
    print(f"  Coincide total cartera:         {coincide_total}")

    # 5) LBs sin CI (potencial bug del Excel)
    if sin_ci:
        print(f"\n========== {len(sin_ci)} LBs con cuota_inicial = 0 ==========")
        for r in sin_ci:
            print(f"  {r['lb']:<14} {(r['cliente'] or '?')[:32]:<32} estado={r['estado']:<10} mo=${r['mo']:>12,} sc+si=${r['sc']+r['si']:>12,} fent={r['fent']}")

    # 6) Generar payload
    OUT = Path(".planning/marcar_cuotas_iniciales_2026-04-30.json")
    OUT.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "fuente":    str(xlsx),
        "generado":  "Day3 B5 — auditoria completa",
        "total_lbs": len(records_ci),
        "total_ci":  total_ci,
        "loanbooks": [
            {k: r[k] for k in ("loanbook_id", "cuota_inicial",
                               "fecha_pago", "metodo_pago")}
            for r in records_ci
        ],
    }
    OUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False),
                   encoding="utf-8")
    print(f"\nGuardado payload (LBs con CI>0): {OUT}")


if __name__ == "__main__":
    main()
