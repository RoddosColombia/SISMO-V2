"""
Genera .planning/marcar_cuotas_iniciales_2026-04-30.json desde Excel oficial.

Uso (desde C:\\Users\\AndresSanJuan\\roddos-workspace\\SISMO-V2):
    python scripts_local/generar_marcar_cuotas_iniciales.py

Output:
    1. Resumen por terminal: 34 LBs (?), Σ CI Excel vs $47.83M
    2. .planning/marcar_cuotas_iniciales_2026-04-30.json — body para el endpoint

Regla: Excel es la única fuente de verdad. Este script solo LEE el Excel.
"""
import json
import sys
from pathlib import Path

import openpyxl

XLSX = Path("loanbook_roddos_2026-04-30.xlsx")
OUT  = Path(".planning/marcar_cuotas_iniciales_2026-04-30.json")

ANDRES_TOTAL_CI  = 47_830_000
ANDRES_TOTAL_LBS = 34


def fecha_iso(v):
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    return str(v) if v else None


def main():
    if not XLSX.exists():
        print(f"ERROR: no encontré {XLSX} en el directorio actual.")
        print(f"  cwd: {Path.cwd()}")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX, data_only=True)

    records = []
    total_ci = 0

    # Loan Tape RDX
    if "Loan Tape RDX" in wb.sheetnames:
        ws = wb["Loan Tape RDX"]
        h = [c.value for c in ws[1]]
        ix = {k: h.index(k) for k in
              ("loanbook_codigo", "cliente_nombre", "moto_modelo",
               "cuota_inicial", "fecha_entrega", "estado")}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[ix["loanbook_codigo"]]:
                continue
            ci = int(row[ix["cuota_inicial"]] or 0)
            if ci <= 0:
                continue
            records.append({
                "loanbook_id":   row[ix["loanbook_codigo"]],
                "cliente":       row[ix["cliente_nombre"]],
                "modelo":        row[ix["moto_modelo"]],
                "cuota_inicial": ci,
                "fecha_pago":    fecha_iso(row[ix["fecha_entrega"]]),
                "metodo_pago":   "cuota_inicial_pre_entrega",
                "sheet":         "RDX",
                "estado_excel":  row[ix["estado"]],
            })
            total_ci += ci

    # Loan Tape RODANTE
    if "Loan Tape RODANTE" in wb.sheetnames:
        ws2 = wb["Loan Tape RODANTE"]
        h2 = [c.value for c in ws2[1]]
        ix2 = {k: h2.index(k) for k in
               ("loanbook_codigo", "cliente_nombre", "producto",
                "cuota_inicial", "fecha_entrega", "estado")}
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if not row[ix2["loanbook_codigo"]]:
                continue
            ci = int(row[ix2["cuota_inicial"]] or 0)
            if ci <= 0:
                continue
            records.append({
                "loanbook_id":   row[ix2["loanbook_codigo"]],
                "cliente":       row[ix2["cliente_nombre"]],
                "modelo":        row[ix2["producto"]],
                "cuota_inicial": ci,
                "fecha_pago":    fecha_iso(row[ix2["fecha_entrega"]]),
                "metodo_pago":   "cuota_inicial_pre_entrega",
                "sheet":         "RODANTE",
                "estado_excel":  row[ix2["estado"]],
            })
            total_ci += ci

    print("=== Reconciliación Excel oficial ===")
    print(f"  LBs con cuota_inicial > 0:   {len(records)}")
    print(f"  Sigma cuota_inicial Excel:   ${total_ci:,}")
    print(f"  Andres confirmo:             ${ANDRES_TOTAL_CI:,}  ({ANDRES_TOTAL_LBS} LBs)")
    coincide = (total_ci == ANDRES_TOTAL_CI) and (len(records) == ANDRES_TOTAL_LBS)
    print(f"  Coincide con Andres:         {coincide}")
    print()

    print("=== Detalle por LB ordenado ===")
    for r in sorted(records, key=lambda x: x["loanbook_id"]):
        cliente = (r["cliente"] or "?")[:32]
        modelo  = (r["modelo"] or "?")[:14]
        print(f"  {r['loanbook_id']:<14} {cliente:<32} "
              f"{modelo:<14} ci=${r['cuota_inicial']:>10,} "
              f"fent={r['fecha_pago']}  [{r['sheet']}]  estado_excel={r['estado_excel']}")

    payload = {
        "fuente":    str(XLSX),
        "generado":  "Day3 B5 — 2026-05-04",
        "total_lbs": len(records),
        "total_ci":  total_ci,
        "loanbooks": [
            {k: r[k] for k in ("loanbook_id", "cuota_inicial",
                               "fecha_pago", "metodo_pago")}
            for r in records
        ],
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False),
                   encoding="utf-8")
    print(f"\nGuardado: {OUT}")

    if not coincide:
        print()
        print("ATENCION: el Excel NO coincide con la cifra de Andres ($47.83M / 34 LBs).")
        print("  Verificar antes de continuar con el endpoint.")
        sys.exit(2)


if __name__ == "__main__":
    main()
